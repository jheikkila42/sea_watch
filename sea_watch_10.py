# -*- coding: utf-8 -*-
"""
STCW-yhteensopiva työvuorogeneraattori
Versio 16.1: Blokkipohjainen lähestymistapa + jatkuvan yön tuki

VAIHE 0: Analysoi jatkuvat yövuorot etukäteen
VAIHE 1: Pakolliset (tulo, lähtö, slussi, shiftaus, op 08-17 ulkopuolella)
VAIHE 2: Laske tarvittavat lisätunnit per dayman
VAIHE 3: Jaa työblokit
VAIHE 4: Validointi
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Värit
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN = PatternFill("solid", fgColor="90EE90")
BLUE = PatternFill("solid", fgColor="ADD8E6")
ORANGE = PatternFill("solid", fgColor="FFA500")
PURPLE = PatternFill("solid", fgColor="C9A0DC")
PINK = PatternFill("solid", fgColor="FFB6C1")
WHITE = PatternFill("solid", fgColor="FFFFFF")

# Vakiot
NORMAL_START = 16   # 08:00 (slotti)
NORMAL_END = 34     # 17:00 (slotti)
LUNCH_START = 23    # 11:30
LUNCH_END = 24      # 12:00
MIN_HOURS = 8
MAX_HOURS = 10


def time_to_slot(h, m=0):
    """Muuntaa ajan slotiksi."""
    return h * 2 + (1 if m >= 30 else 0)


def slot_to_time_str(slot):
    """Muuntaa slotin ajaksi."""
    h = slot // 2
    m = "30" if slot % 2 else "00"
    return f"{h:02d}:{m}"


def get_work_ranges(work_slots):
    """Palauttaa työjaksot luettavassa muodossa."""
    ranges = []
    start = None
    for i, w in enumerate(work_slots):
        if w and start is None:
            start = i
        elif not w and start is not None:
            ranges.append(f"{slot_to_time_str(start)}-{slot_to_time_str(i)}")
            start = None
    if start is not None:
        ranges.append(f"{slot_to_time_str(start)}-00:00")
    return ranges


# ============================================================================
# STCW-TARKISTUS
# ============================================================================

def analyze_stcw(work_48h):
    """
    Analysoi STCW-lepoajat 48h (2 päivän) työvuorolistasta.
    Yhdistää yölevon jos se jatkuu päivien yli.
    """
    rest_slots = [not w for w in work_48h[:48]]
    
    # Laske lepojaksot (vähintään 1h)
    rest_periods = []
    current_rest = 0
    for is_rest in rest_slots:
        if is_rest:
            current_rest += 1
        else:
            if current_rest >= 2:  # Min 1h
                rest_periods.append(current_rest / 2)
            current_rest = 0
    if current_rest >= 2:
        rest_periods.append(current_rest / 2)
    
    # Yhdistä yölepo jos se jatkuu päivien yli (keskiyön yli)
    if len(rest_periods) >= 2 and len(work_48h) >= 48:
        if not work_48h[47] and not work_48h[0]:
            # Lepo jatkuu keskiyön yli - yhdistä ensimmäinen ja viimeinen
            combined = rest_periods[-1] + rest_periods[0]
            rest_periods = [combined] + rest_periods[1:-1]
    
    total_rest = sum(rest_periods)
    longest_rest = max(rest_periods) if rest_periods else 0
    
    # Kaksi pisintä jaksoa
    sorted_periods = sorted(rest_periods, reverse=True)
    top_two = sorted_periods[:2] if len(sorted_periods) >= 2 else sorted_periods
    top_two_total = sum(top_two)
    
    issues = []
    if top_two_total < 10:
        issues.append(f"Lepoa vain {top_two_total}h (min 10h kahdessa jaksossa)")
    if longest_rest < 6:
        issues.append(f"Pisin lepo {longest_rest}h (min 6h)")
    
    return {
        'total_rest': total_rest,
        'rest_period_count': len(rest_periods),
        'longest_rest': longest_rest,
        'top_two_total': top_two_total,
        'status': 'OK' if not issues else 'RIKE',
        'issues': issues
    }


def check_stcw(work_day1, work_day2):
    """
    Tarkistaa STCW-säännöt kahden päivän välillä.
    """
    combined = work_day1 + work_day2
    ana = analyze_stcw(combined)
    
    return {
        'ok': ana['status'] == 'OK',
        'rest_periods': ana['rest_period_count'],
        'total_rest': ana['total_rest'],
        'longest_rest': ana['longest_rest'],
        'issues': ana['issues']
    }


def check_stcw_ok(work_slots, prev_day_work=None):
    """
    Tarkistaa onko STCW OK 24h ikkunassa päivän lopussa.
    """
    if prev_day_work is None:
        prev_day_work = [False] * 48
    
    combined = prev_day_work + work_slots
    ana = analyze_stcw(combined)
    
    return ana['status'] == 'OK'


def check_stcw_at_slot(work_96h, slot_index):
    """
    Tarkistaa STCW tietyssä slotissa (app.py yhteensopivuus).
    """
    start = max(0, slot_index - 47)
    end = slot_index + 1
    window = work_96h[start:end]
    
    if len(window) < 48:
        padding = [False] * (48 - len(window))
        window = padding + window
    
    ana = analyze_stcw(window + [False] * 48)
    
    return {
        'total_rest': ana['total_rest'],
        'longest_rest': ana['longest_rest'],
        'rest_period_count': ana['rest_period_count'],
        'status': ana['status']
    }


# ============================================================================
# APUFUNKTIOT
# ============================================================================

def get_work_blocks(work_slots):
    """Palauttaa työblokit listana (start_slot, end_slot)."""
    blocks = []
    start = None
    
    for i, w in enumerate(work_slots):
        if w and start is None:
            start = i
        elif not w and start is not None:
            blocks.append((start, i))
            start = None
    
    if start is not None:
        blocks.append((start, len(work_slots)))
    
    return blocks


def add_block(work_slots, start, end, marker_slots=None):
    """Lisää työblokki ja merkitsee sen haluttuun marker-listaan."""
    for i in range(max(0, start), min(end, 48)):
        work_slots[i] = True
        if marker_slots is not None:
            marker_slots[i] = True


def evaluate_night_split(prev_early, prev_late, split_slot, arrival_start=None, 
                          departure_start=None):
    """
    Arvioi yövuoron jakokohdan sopivuus STCW:n kannalta.
    """
    early_work = [False] * 48
    late_work = [False] * 48
    
    # Early worker: 00:00 -> split_slot
    for slot in range(0, min(split_slot, 48)):
        early_work[slot] = True
    
    # Late worker: split_slot -> 08:00
    for slot in range(split_slot, min(NORMAL_START, 48)):
        late_work[slot] = True
    
    # Lisää tulo/lähtö molemmille jos on
    if arrival_start is not None:
        for i in range(arrival_start, min(arrival_start + 2, 48)):
            early_work[i] = True
            late_work[i] = True
    if departure_start is not None:
        for i in range(departure_start, min(departure_start + 2, 48)):
            early_work[i] = True
            late_work[i] = True
    
    early_ana = analyze_stcw(prev_early + early_work)
    late_ana = analyze_stcw(prev_late + late_work)
    
    early_issues = len(early_ana['issues'])
    late_issues = len(late_ana['issues'])
    total_issues = early_issues + late_issues
    
    min_longest_rest = min(early_ana['longest_rest'], late_ana['longest_rest'])
    min_total_rest = min(early_ana['top_two_total'], late_ana['top_two_total'])
    
    # Pienempi score on parempi
    return (total_issues, -min_longest_rest, -min_total_rest)


def choose_night_split_slot(prev_early, prev_late, arrival_start=None, 
                            departure_start=None):
    """
    Valitsee optimaalisen yövuoron jakokohdan (01:00 - 07:00 väliltä).
    """
    candidate_slots = list(range(time_to_slot(1, 0), time_to_slot(7, 0) + 1))
    best_slot = time_to_slot(3, 0)  # Oletus 03:00
    best_score = None
    
    for split_slot in candidate_slots:
        score = evaluate_night_split(prev_early, prev_late, split_slot, 
                                     arrival_start, departure_start)
        if best_score is None or score < best_score:
            best_score = score
            best_slot = split_slot
    
    return best_slot


# ============================================================================
# PÄÄFUNKTIO
# ============================================================================

def generate_schedule(days_data):
    """
    Generoi työvuorot blokkipohjaisella lähestymistavalla.
    """
    workers = ['Bosun', 'Dayman EU', 'Dayman PH1', 'Dayman PH2',
               'Watchman 1', 'Watchman 2', 'Watchman 3']
    daymen = ['Dayman EU', 'Dayman PH1', 'Dayman PH2']
    
    all_days = {w: [] for w in workers}
    num_days = len(days_data)
    
    # ========================================================================
    # VAIHE 0: ANALYSOI JATKUVAT YÖVUOROT ETUKÄTEEN
    # ========================================================================
    
    continuous_nights = []
    
    for d in range(num_days - 1):
        curr = days_data[d]
        next_day = days_data[d + 1]
        
        curr_op_end = curr.get('port_op_end_hour')
        next_op_start = next_day.get('port_op_start_hour')
        
        # Jos päivä N loppuu 00:00 ja päivä N+1 alkaa 00:00 -> jatkuva yö
        if curr_op_end == 0 and next_op_start == 0:
            continuous_nights.append({
                'day_index': d,
                'early_worker': 'Dayman PH1',  # Tekee alkuyön
                'late_worker': 'Dayman PH2'    # Tekee loppuyön
            })
    
    # ========================================================================
    # GENEROI PÄIVÄ KERRALLAAN
    # ========================================================================
    
    for d, info in enumerate(days_data):
        
        # Parsitaan ajat
        op_start_h = info.get('port_op_start_hour')
        op_start_m = info.get('port_op_start_minute', 0)
        op_end_h = info.get('port_op_end_hour')
        op_end_m = info.get('port_op_end_minute', 0)
        
        arrival_h = info.get('arrival_hour')
        arrival_m = info.get('arrival_minute', 0)
        departure_h = info.get('departure_hour')
        departure_m = info.get('departure_minute', 0)
        
        sluice_arr_h = info.get('sluice_arrival_hour')
        sluice_arr_m = info.get('sluice_arrival_minute', 0)
        sluice_dep_h = info.get('sluice_departure_hour')
        sluice_dep_m = info.get('sluice_departure_minute', 0)
        shifting_h = info.get('shifting_hour')
        shifting_m = info.get('shifting_minute', 0)
        
        # Muunna sloiteiksi
        if op_start_h is not None:
            op_start = time_to_slot(op_start_h, op_start_m)
            if op_end_h is not None and op_end_h < op_start_h:
                op_end = 48  # Yli keskiyön
            elif op_end_h == 0 and op_start_h > 0:
                op_end = 48
            elif op_end_h is not None:
                op_end = time_to_slot(op_end_h, op_end_m)
            else:
                op_end = NORMAL_END
        else:
            op_start = NORMAL_START
            op_end = NORMAL_END
        
        arrival_start = time_to_slot(arrival_h, arrival_m) if arrival_h is not None else None
        departure_start = time_to_slot(departure_h, departure_m) if departure_h is not None else None
        sluice_arr_start = time_to_slot(sluice_arr_h, sluice_arr_m) if sluice_arr_h is not None else None
        sluice_dep_start = time_to_slot(sluice_dep_h, sluice_dep_m) if sluice_dep_h is not None else None
        shifting_start = time_to_slot(shifting_h, shifting_m) if shifting_h is not None else None
        
        # Tarkista onko tämä päivä jatkuvan yön jälkeen
        continues_from_night = False
        night_split_slot = None
        early_worker = None
        late_worker = None
        
        for night_info in continuous_nights:
            if night_info['day_index'] == d - 1:
                continues_from_night = True
                early_worker = night_info['early_worker']
                late_worker = night_info['late_worker']
                
                # Laske optimaalinen jakokohta
                prev_early = all_days[early_worker][d - 1]['work_slots'] if d > 0 else [False] * 48
                prev_late = all_days[late_worker][d - 1]['work_slots'] if d > 0 else [False] * 48
                night_split_slot = choose_night_split_slot(prev_early, prev_late, 
                                                           arrival_start, departure_start)
                break
        
        # Tarkista aloittaako tämä päivä jatkuvan yön
        starts_night = False
        for night_info in continuous_nights:
            if night_info['day_index'] == d:
                starts_night = True
                break
        
        # Edellisen päivän työvuorot (STCW-tarkistukseen)
        prev_day_work = {}
        for dm in daymen:
            if d > 0:
                prev_day_work[dm] = all_days[dm][d - 1]['work_slots']
            else:
                prev_day_work[dm] = [False] * 48
        
        # Alusta työntekijöiden data
        dm_work = {dm: [False] * 48 for dm in daymen}
        dm_arr = {dm: [False] * 48 for dm in daymen}
        dm_dep = {dm: [False] * 48 for dm in daymen}
        dm_ops = {dm: [False] * 48 for dm in daymen}
        dm_sluice = {dm: [False] * 48 for dm in daymen}
        dm_shifting = {dm: [False] * 48 for dm in daymen}
        
        # ====================================================================
        # VAIHE 1: PAKOLLISET
        # ====================================================================
        
        # 1.1: Tulo - kaikki daymanit (1h)
        if arrival_start is not None:
            for dm in daymen:
                add_block(dm_work[dm], arrival_start, arrival_start + 2, dm_arr[dm])
        
        # 1.2: Lähtö - 2 daymaniä (1h)
        if departure_start is not None:
            scores = {}
            for dm in daymen:
                hours = sum(dm_work[dm]) / 2
                continuity = 1 if (departure_start > 0 and dm_work[dm][departure_start - 1]) else 0
                scores[dm] = -hours + continuity
            
            selected = sorted(daymen, key=lambda x: scores[x], reverse=True)[:2]
            for dm in selected:
                add_block(dm_work[dm], departure_start, departure_start + 2, dm_dep[dm])
        
        # 1.3: Slussi tulo - 1. tunti 2 dm, 2. tunti 3 dm (2h kokonaan)
        if sluice_arr_start is not None:
            scores = {}
            for dm in daymen:
                hours = sum(dm_work[dm]) / 2
                scores[dm] = -hours
            
            first_hour_dm = sorted(daymen, key=lambda x: scores[x], reverse=True)[:2]
            
            for dm in first_hour_dm:
                add_block(dm_work[dm], sluice_arr_start, sluice_arr_start + 2, dm_sluice[dm])
            
            for dm in daymen:
                add_block(dm_work[dm], sluice_arr_start + 2, sluice_arr_start + 4, dm_sluice[dm])
        
        # 1.4: Slussi lähtö - 2 daymaniä (2h)
        if sluice_dep_start is not None:
            scores = {}
            for dm in daymen:
                hours = sum(dm_work[dm]) / 2
                continuity = 1 if (sluice_dep_start > 0 and dm_work[dm][sluice_dep_start - 1]) else 0
                scores[dm] = -hours + continuity
            
            selected = sorted(daymen, key=lambda x: scores[x], reverse=True)[:2]
            for dm in selected:
                add_block(dm_work[dm], sluice_dep_start, sluice_dep_start + 4, dm_sluice[dm])
        
        # 1.5: Shiftaus - kaikki daymanit (1h)
        if shifting_start is not None:
            for dm in daymen:
                add_block(dm_work[dm], shifting_start, shifting_start + 2, dm_shifting[dm])
        
        # 1.6: Satamaop 08-17 ULKOPUOLELLA - aina 1 dayman töissä
        # Käsitellään SLOTTI KERRALLAAN jatkuvuuden varmistamiseksi
        op_outside_slots = []
        for slot in range(op_start, min(op_end, 48)):
            if slot < NORMAL_START or slot >= NORMAL_END:
                if slot != LUNCH_START:
                    op_outside_slots.append(slot)
        
        current_worker = None
        
        # Jos jatkuva yö edellisestä päivästä, käytä jakoa
        if continues_from_night and night_split_slot is not None:
            # Early worker tekee 00:00 -> split
            for slot in range(0, night_split_slot):
                if slot in op_outside_slots:
                    dm_work[early_worker][slot] = True
                    dm_ops[early_worker][slot] = True
            
            # Late worker tekee split -> 08:00
            for slot in range(night_split_slot, NORMAL_START):
                if slot in op_outside_slots:
                    dm_work[late_worker][slot] = True
                    dm_ops[late_worker][slot] = True
            
            # Poista nämä slotit op_outside_slots:ista
            op_outside_slots = [s for s in op_outside_slots if s >= NORMAL_START]
        
        for slot in op_outside_slots:
            can_continue = False
            
            if current_worker is not None:
                current_hours = sum(dm_work[current_worker]) / 2
                
                if current_hours < MAX_HOURS:
                    test_work = dm_work[current_worker][:]
                    test_work[slot] = True
                    stcw_ok = check_stcw_ok(test_work, prev_day_work[current_worker])
                    
                    if stcw_ok:
                        can_continue = True
            
            if can_continue:
                dm_work[current_worker][slot] = True
                dm_ops[current_worker][slot] = True
            else:
                best_dm = None
                best_score = -9999
                
                for dm in daymen:
                    if dm == current_worker:
                        continue
                    
                    current_hours = sum(dm_work[dm]) / 2
                    
                    if current_hours >= MAX_HOURS:
                        continue
                    
                    test_work = dm_work[dm][:]
                    test_work[slot] = True
                    stcw_ok = check_stcw_ok(test_work, prev_day_work[dm])
                    
                    if not stcw_ok:
                        continue
                    
                    score = 0
                    score += (MAX_HOURS - current_hours) * 10
                    
                    # Priorisoi daymaniä jolla pidempi lepo edellisestä päivästä
                    prev_work = prev_day_work[dm]
                    last_work_slot = -1
                    for s in range(47, -1, -1):
                        if prev_work[s]:
                            last_work_slot = s
                            break
                    
                    if last_work_slot >= 0:
                        rest_slots = (slot + 48) - last_work_slot
                        rest_hours = rest_slots / 2
                        score += min(rest_hours, 24) * 5
                    else:
                        score += 24 * 5
                    
                    if score > best_score:
                        best_score = score
                        best_dm = dm
                
                if best_dm:
                    dm_work[best_dm][slot] = True
                    dm_ops[best_dm][slot] = True
                    current_worker = best_dm
                else:
                    if current_worker is not None:
                        current_hours = sum(dm_work[current_worker]) / 2
                        if current_hours < MAX_HOURS:
                            dm_work[current_worker][slot] = True
                            dm_ops[current_worker][slot] = True
        
        # ====================================================================
        # VAIHE 2: LASKE TARVITTAVAT LISÄTUNNIT
        # ====================================================================
        
        needed_hours = {}
        for dm in daymen:
            current = sum(dm_work[dm]) / 2
            needed_hours[dm] = max(0, MIN_HOURS - current)
        
        # ====================================================================
        # VAIHE 3: JAA TYÖBLOKIT
        # ====================================================================
        
        # 3.1: Op-kattavuus 08-17 välillä
        op_inside_slots = []
        for slot in range(max(op_start, NORMAL_START), min(op_end, NORMAL_END)):
            if LUNCH_START <= slot < LUNCH_END:
                continue
            op_inside_slots.append(slot)
        
        for slot in op_inside_slots:
            workers_in_slot = [dm for dm in daymen if dm_work[dm][slot]]
            
            if len(workers_in_slot) >= 1:
                for dm in workers_in_slot:
                    dm_ops[dm][slot] = True
                continue
            
            best_dm = None
            best_score = -9999
            
            for dm in daymen:
                current_hours = sum(dm_work[dm]) / 2
                
                if current_hours >= MAX_HOURS:
                    continue
                
                score = 0
                
                if slot > 0 and dm_work[dm][slot - 1]:
                    score += 200
                if slot < 47 and dm_work[dm][slot + 1]:
                    score += 200
                
                test_work = dm_work[dm][:]
                test_work[slot] = True
                stcw_ok = check_stcw_ok(test_work, prev_day_work[dm])
                
                if not stcw_ok:
                    continue
                
                score += (MAX_HOURS - current_hours) * 10
                
                if score > best_score:
                    best_score = score
                    best_dm = dm
            
            if best_dm:
                dm_work[best_dm][slot] = True
                dm_ops[best_dm][slot] = True
        
        # 3.2: Täytä loput tunnit 08:00 alkaen
        for dm in daymen:
            current_hours = sum(dm_work[dm]) / 2
            
            if current_hours >= MIN_HOURS:
                continue
            
            night_work_slots = sum(1 for s in range(0, NORMAL_START) if dm_work[dm][s])
            did_night_shift = night_work_slots >= 4
            
            if did_night_shift:
                last_night_slot = -1
                for s in range(NORMAL_START - 1, -1, -1):
                    if dm_work[dm][s]:
                        last_night_slot = s
                        break
                
                if last_night_slot >= 0 and last_night_slot < NORMAL_START - 1:
                    for s in range(last_night_slot + 1, NORMAL_START):
                        if current_hours >= MIN_HOURS:
                            break
                        dm_work[dm][s] = True
                        if op_start <= s < min(op_end, 48):
                            dm_ops[dm][s] = True
                        current_hours = sum(dm_work[dm]) / 2
                
                continue
            
            slot = NORMAL_START
            
            while current_hours < MIN_HOURS and slot < NORMAL_END:
                if LUNCH_START <= slot < LUNCH_END:
                    slot += 1
                    continue
                
                if dm_work[dm][slot]:
                    slot += 1
                    continue
                
                dm_work[dm][slot] = True
                if op_start <= slot < min(op_end, 48):
                    dm_ops[dm][slot] = True
                current_hours = sum(dm_work[dm]) / 2
                slot += 1
        
        # 3.3: Varmista op-kattavuus uudelleen
        for slot in op_inside_slots:
            workers_in_slot = [dm for dm in daymen if dm_work[dm][slot]]
            
            if len(workers_in_slot) >= 1:
                for dm in workers_in_slot:
                    dm_ops[dm][slot] = True
                continue
            
            best_dm = None
            best_score = -9999
            
            for dm in daymen:
                current_hours = sum(dm_work[dm]) / 2
                
                if current_hours >= MAX_HOURS:
                    continue
                
                score = 0
                
                if slot > 0 and dm_work[dm][slot - 1]:
                    score += 200
                if slot < 47 and dm_work[dm][slot + 1]:
                    score += 200
                
                score += (MAX_HOURS - current_hours) * 10
                
                if score > best_score:
                    best_score = score
                    best_dm = dm
            
            if best_dm:
                dm_work[best_dm][slot] = True
                dm_ops[best_dm][slot] = True
        
        # 3.4: Täytä turhat aukot blokkien välissä
        for dm in daymen:
            work = dm_work[dm]
            blocks = get_work_blocks(work)
            
            for i in range(len(blocks) - 1):
                _, block1_end = blocks[i]
                block2_start, _ = blocks[i + 1]
                
                gap = block2_start - block1_end
                
                if 0 < gap <= 4 and block1_end >= NORMAL_START and block2_start <= NORMAL_END:
                    current_hours = sum(work) / 2
                    
                    for s in range(block1_end, block2_start):
                        if LUNCH_START <= s < LUNCH_END:
                            continue
                        if current_hours >= MAX_HOURS:
                            break
                        work[s] = True
                        if op_start <= s < min(op_end, 48):
                            dm_ops[dm][s] = True
                        current_hours = sum(work) / 2
        
        # ====================================================================
        # VAIHE 4: TÄYTÄ AUKOT (max 1h aukot)
        # ====================================================================
        
        for dm in daymen:
            work = dm_work[dm]
            blocks = get_work_blocks(work)
            
            for i in range(len(blocks) - 1):
                _, block1_end = blocks[i]
                block2_start, _ = blocks[i + 1]
                
                gap = block2_start - block1_end
                
                if 0 < gap <= 2:
                    for s in range(block1_end, block2_start):
                        if LUNCH_START <= s < LUNCH_END:
                            continue
                        work[s] = True
                        if op_start <= s < min(op_end, 48):
                            dm_ops[dm][s] = True
        
        # ====================================================================
        # TALLENNA TULOKSET
        # ====================================================================
        
        for dm in daymen:
            all_days[dm].append({
                'work_slots': dm_work[dm],
                'arrival_slots': dm_arr[dm],
                'departure_slots': dm_dep[dm],
                'port_op_slots': dm_ops[dm],
                'sluice_slots': dm_sluice[dm],
                'shifting_slots': dm_shifting[dm]
            })
        
        # ====================================================================
        # BOSUN (08-17 + tulo/lähtö + slussi + shiftaus)
        # ====================================================================
        
        bosun_work = [False] * 48
        bosun_arr = [False] * 48
        bosun_dep = [False] * 48
        bosun_sluice = [False] * 48
        bosun_shifting = [False] * 48
        
        # Tulo
        if arrival_start is not None:
            add_block(bosun_work, arrival_start, arrival_start + 2, bosun_arr)
        
        # Lähtö
        if departure_start is not None:
            add_block(bosun_work, departure_start, departure_start + 2, bosun_dep)
        
        # Slussi tulo (2h)
        if sluice_arr_start is not None:
            add_block(bosun_work, sluice_arr_start, sluice_arr_start + 4, bosun_sluice)
        
        # Slussi lähtö (2h)
        if sluice_dep_start is not None:
            add_block(bosun_work, sluice_dep_start, sluice_dep_start + 4, bosun_sluice)
        
        # Shiftaus (1h)
        if shifting_start is not None:
            add_block(bosun_work, shifting_start, shifting_start + 2, bosun_shifting)
        
        # Täytä loput 08-17 (paitsi lounas)
        for slot in range(NORMAL_START, NORMAL_END):
            if slot != LUNCH_START:
                bosun_work[slot] = True
        
        all_days['Bosun'].append({
            'work_slots': bosun_work,
            'arrival_slots': bosun_arr,
            'departure_slots': bosun_dep,
            'port_op_slots': [False] * 48,
            'sluice_slots': bosun_sluice,
            'shifting_slots': bosun_shifting
        })
        
        # ====================================================================
        # WATCHMANIT (4h vuorot)
        # ====================================================================
        
        watch_schedules = {
            'Watchman 1': [(0, 4), (12, 16)],
            'Watchman 2': [(4, 8), (16, 20)],
            'Watchman 3': [(8, 12), (20, 24)]
        }
        
        for watchman, shifts in watch_schedules.items():
            work = [False] * 48
            for (start_h, end_h) in shifts:
                for i in range(start_h * 2, end_h * 2):
                    work[i] = True
            
            all_days[watchman].append({
                'work_slots': work,
                'arrival_slots': [False] * 48,
                'departure_slots': [False] * 48,
                'port_op_slots': [False] * 48,
                'sluice_slots': [False] * 48,
                'shifting_slots': [False] * 48
            })
    
    # ========================================================================
    # LUO EXCEL JA RAPORTTI
    # ========================================================================
    
    wb = Workbook()
    wb.remove(wb.active)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for d in range(num_days):
        ws = wb.create_sheet(title=f"Päivä {d + 1}")
        
        ws.cell(row=1, column=1, value="Työntekijä")
        for col in range(48):
            cell = ws.cell(row=1, column=col + 2, value=slot_to_time_str(col))
            cell.alignment = Alignment(horizontal='center', textRotation=90)
            cell.font = Font(size=8)
            cell.border = thin_border
        ws.cell(row=1, column=50, value="Tunnit")
        
        for row, worker in enumerate(workers, start=2):
            ws.cell(row=row, column=1, value=worker)
            
            data = all_days[worker][d]
            work = data['work_slots']
            arr = data['arrival_slots']
            dep = data['departure_slots']
            ops = data['port_op_slots']
            sluice = data.get('sluice_slots', [False] * 48)
            shifting = data.get('shifting_slots', [False] * 48)
            
            hours = sum(work) / 2
            
            for col in range(48):
                cell = ws.cell(row=row, column=col + 2)
                cell.border = thin_border
                
                if work[col]:
                    if sluice[col]:
                        cell.fill = PURPLE
                        cell.value = "SL"
                    elif shifting[col]:
                        cell.fill = PINK
                        cell.value = "SH"
                    elif arr[col]:
                        cell.fill = GREEN
                        cell.value = "B"
                    elif dep[col]:
                        cell.fill = BLUE
                        cell.value = "C"
                    elif ops[col]:
                        cell.fill = YELLOW
                        cell.value = "OP"
                    else:
                        cell.fill = ORANGE
                        cell.value = "X"
                else:
                    cell.fill = WHITE
            
            ws.cell(row=row, column=50, value=f"{hours}h")
        
        ws.column_dimensions['A'].width = 12
        for col in range(2, 50):
            ws.column_dimensions[get_column_letter(col)].width = 3.5
        ws.column_dimensions[get_column_letter(50)].width = 6
    
    # STCW-raportti
    report = []
    for d in range(1, num_days):
        for w in workers:
            work1 = all_days[w][d - 1]['work_slots']
            work2 = all_days[w][d]['work_slots']
            stcw_result = check_stcw(work1, work2)
            report.append({
                'day': d + 1,
                'worker': w,
                'stcw': stcw_result
            })
    
    return wb, all_days, report


def build_workbook_and_report(all_days, num_days, workers):
    """
    Rakentaa Excelin ja raportin valmiista all_days-datasta.
    (app.py yhteensopivuus muokkauksen jälkeen)
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for d in range(num_days):
        ws = wb.create_sheet(title=f"Päivä {d + 1}")
        
        ws.cell(row=1, column=1, value="Työntekijä")
        for col in range(48):
            cell = ws.cell(row=1, column=col + 2, value=slot_to_time_str(col))
            cell.alignment = Alignment(horizontal='center', textRotation=90)
            cell.font = Font(size=8)
            cell.border = thin_border
        ws.cell(row=1, column=50, value="Tunnit")
        
        for row, worker in enumerate(workers, start=2):
            ws.cell(row=row, column=1, value=worker)
            
            data = all_days[worker][d]
            work = data['work_slots']
            arr = data['arrival_slots']
            dep = data['departure_slots']
            ops = data['port_op_slots']
            sluice = data.get('sluice_slots', [False] * 48)
            shifting = data.get('shifting_slots', [False] * 48)
            
            hours = sum(work) / 2
            
            for col in range(48):
                cell = ws.cell(row=row, column=col + 2)
                cell.border = thin_border
                
                if work[col]:
                    if sluice[col]:
                        cell.fill = PURPLE
                        cell.value = "SL"
                    elif shifting[col]:
                        cell.fill = PINK
                        cell.value = "SH"
                    elif arr[col]:
                        cell.fill = GREEN
                        cell.value = "B"
                    elif dep[col]:
                        cell.fill = BLUE
                        cell.value = "C"
                    elif ops[col]:
                        cell.fill = YELLOW
                        cell.value = "OP"
                    else:
                        cell.fill = ORANGE
                        cell.value = "X"
                else:
                    cell.fill = WHITE
            
            ws.cell(row=row, column=50, value=f"{hours}h")
        
        ws.column_dimensions['A'].width = 12
        for col in range(2, 50):
            ws.column_dimensions[get_column_letter(col)].width = 3.5
        ws.column_dimensions[get_column_letter(50)].width = 6
    
    report = []
    return wb, report


# ============================================================================
# TESTAUS
# ============================================================================

if __name__ == "__main__":
    print("Generoidaan työvuorot (blokkipohjainen)...")
    
    days_data = [
        {
            'arrival_hour': 18, 'arrival_minute': 0,
            'departure_hour': None, 'departure_minute': 0,
            'port_op_start_hour': 19, 'port_op_start_minute': 0,
            'port_op_end_hour': 0, 'port_op_end_minute': 0
        },
        {
            'arrival_hour': None, 'arrival_minute': 0,
            'departure_hour': 20, 'departure_minute': 0,
            'port_op_start_hour': 0, 'port_op_start_minute': 0,
            'port_op_end_hour': 19, 'port_op_end_minute': 0
        }
    ]
    
    wb, all_days, report = generate_schedule(days_data)
    
    print("\n" + "=" * 60)
    print("TULOKSET")
    print("=" * 60)
    
    for d in range(len(days_data)):
        info = days_data[d]
        arr = f"{info['arrival_hour']:02d}:00" if info['arrival_hour'] else "-"
        dep = f"{info['departure_hour']:02d}:00" if info['departure_hour'] else "-"
        
        print(f"\n=== Päivä {d + 1} ===")
        print(f"  Tulo: {arr} | Lähtö: {dep}")
        
        for w in ['Dayman EU', 'Dayman PH1', 'Dayman PH2']:
            work = all_days[w][d]['work_slots']
            hours = sum(work) / 2
            ranges = get_work_ranges(work)
            print(f"\n  {w}: {hours}h | {' + '.join(ranges)}")
    
    print("\n" + "=" * 60)
    print("STCW-TARKISTUS")
    print("=" * 60)
    
    for w in ['Dayman EU', 'Dayman PH1', 'Dayman PH2']:
        if len(all_days[w]) >= 2:
            work1 = all_days[w][0]['work_slots']
            work2 = all_days[w][1]['work_slots']
            result = check_stcw(work1, work2)
            
            status = "✓" if result['ok'] else "⚠"
            print(f"\n{w}: {status}")
            print(f"  Lepoa: {result['total_rest']}h | Pisin: {result['longest_rest']}h | Jaksoja: {result['rest_periods']}")
            if result['issues']:
                for issue in result['issues']:
                    print(f"  ⚠ {issue}")
    
    wb.save("tyovuorot_v16.xlsx")
    print(f"\nExcel tallennettu: tyovuorot_v16.xlsx")
