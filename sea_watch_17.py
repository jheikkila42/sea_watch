# -*- coding: utf-8 -*-
"""
STCW-yhteensopiva työvuorogeneraattori
Versio 17: Pilkottu ja refaktoroitu

Vaiheet:
- VAIHE 0: Analysoi jatkuvat yövuorot etukäteen
- VAIHE 1: Pakolliset (tulo, lähtö, slussi, shiftaus, op 08-17 ulkopuolella)
- VAIHE 2: Laske tarvittavat lisätunnit per dayman
- VAIHE 3: Jaa työblokit
- VAIHE 4: Täytä aukot
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ============================================================================
# VAKIOT
# ============================================================================

YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN = PatternFill("solid", fgColor="90EE90")
BLUE = PatternFill("solid", fgColor="ADD8E6")
ORANGE = PatternFill("solid", fgColor="FFA500")
PURPLE = PatternFill("solid", fgColor="C9A0DC")
PINK = PatternFill("solid", fgColor="FFB6C1")
WHITE = PatternFill("solid", fgColor="FFFFFF")

NORMAL_START = 16   # 08:00 (slotti)
NORMAL_END = 34     # 17:00 (slotti)
LUNCH_START = 23    # 11:30
LUNCH_END = 24      # 12:00
MIN_HOURS = 8
MAX_HOURS = 10

WORKERS = ['Bosun', 'Dayman EU', 'Dayman PH1', 'Dayman PH2',
           'Watchman 1', 'Watchman 2', 'Watchman 3']
DAYMEN = ['Dayman EU', 'Dayman PH1', 'Dayman PH2']
WATCHMEN = ['Watchman 1', 'Watchman 2', 'Watchman 3']


def add_time_cell_conditional_formatting(ws, start_row, end_row):
    """Lisää työaikasoluille ehdolliset värit solun arvon perusteella."""
    if start_row > end_row:
        return

    time_range = f"B{start_row}:AW{end_row}"
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"SL"'], fill=PURPLE))
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"SH"'], fill=PINK))
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"B"'], fill=YELLOW))
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"C"'], fill=ORANGE))
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"OP"'], fill=GREEN))
    ws.conditional_formatting.add(time_range, CellIsRule(operator='equal', formula=['"X"'], fill=BLUE))


# ============================================================================
# AIKA- JA SLOTTIFUNKTIOT
# ============================================================================

def time_to_slot(h, m=0):
    """Muuntaa ajan slotiksi (0-47)."""
    return h * 2 + (1 if m >= 30 else 0)


def slot_to_time_str(slot):
    """Muuntaa slotin aikamerkkijonoksi (HH:MM)."""
    h = slot // 2
    m = "30" if slot % 2 else "00"
    return f"{h:02d}:{m}"


def get_work_ranges(work_slots):
    """Palauttaa työjaksot luettavassa muodossa."""
    ranges = []
    start = None
    for i, w in enumerate(work_slots):
        if w and start is None:
            start = i
        elif not w and start is not None:
            ranges.append(f"{slot_to_time_str(start)}-{slot_to_time_str(i)}")
            start = None
    if start is not None:
        ranges.append(f"{slot_to_time_str(start)}-00:00")
    return ranges


def add_block(work_list, start, end, marker_list=None):
    """
    Lisää työblokin slottilistaan.
    
    Returns:
        overflow: Montako slottia meni yli päivän (0 jos ei ylivuotoa)
    """
    overflow = 0
    for i in range(start, end):
        if i < 48:
            work_list[i] = True
            if marker_list is not None:
                marker_list[i] = True
        else:
            overflow += 1
    return overflow


def get_work_blocks(work_slots):
    """Palauttaa työblokit listana (start, end) -tupleja."""
    blocks = []
    start = None
    for i, w in enumerate(work_slots):
        if w and start is None:
            start = i
        elif not w and start is not None:
            blocks.append((start, i))
            start = None
    if start is not None:
        blocks.append((start, 48))
    return blocks


def parse_time_str(time_str):
    """Parsii aikamerkkijonon (HH:MM) slotiksi."""
    if not time_str:
        return None
    if isinstance(time_str, int):
        return time_str * 2
    time_str = str(time_str).replace(".", ":")
    parts = time_str.split(":")
    h = int(parts[0])
    m = int(parts[1]) if len(parts) > 1 else 0
    return time_to_slot(h, m)


# ============================================================================
# WATCHMAN-FUNKTIOT
# ============================================================================

# Watchmanien vahtivuorot ja laajennusrajat
WATCHMAN_SHIFTS = {
    'Watchman 1': {
        'shifts': [(0, 8), (24, 32)],      # 00-04, 12-16
        'early_start': 44,                  # 22:00 (2h ennen 00:00)
        'late_end': 36                      # 18:00 (2h jälkeen 16:00)
    },
    'Watchman 2': {
        'shifts': [(8, 16), (32, 40)],     # 04-08, 16-20
        'early_start': 4,                   # 02:00 (2h ennen 04:00)
        'late_end': 44                      # 22:00 (2h jälkeen 20:00)
    },
    'Watchman 3': {
        'shifts': [(16, 24), (40, 48)],    # 08-12, 20-24
        'early_start': 12,                  # 06:00 (2h ennen 08:00)
        'late_end': 52                      # 02:00 seuraavana päivänä (2h jälkeen 00:00)
    }
}


def is_within_watchman_shift(watchman, slot):
    """
    Tarkistaa onko slotti watchmanin normaalin vahtivuoron sisällä.
    """
    shifts = WATCHMAN_SHIFTS.get(watchman, {}).get('shifts', [])
    for start, end in shifts:
        if start <= slot < end:
            return True
    return False


def can_watchman_take_slot(watchman, slot, watchman_states):
    """
    Tarkistaa voiko watchman ottaa tämän slotin.
    
    Watchman voi työskennellä:
    - Oman vahtivuoronsa aikana (aina OK)
    - 2h ennen ensimmäistä vahtia TAI 2h jälkeen viimeistä vahtia
    - Ei molempia samana päivänä
    
    Returns:
        True jos watchman voi ottaa slotin
    """
    # Vahtivuoron sisällä - aina OK
    if is_within_watchman_shift(watchman, slot):
        return True
    
    state = watchman_states.get(watchman, {'extended_start': False, 'extended_end': False})
    shift_info = WATCHMAN_SHIFTS.get(watchman, {})
    
    early_start = shift_info.get('early_start', 0)
    late_end = shift_info.get('late_end', 48)
    
    # Watchman 1: early_start=44 (22:00), ensimmäinen vahti alkaa 0 (00:00)
    # Slotti on "aikaisessa laajennuksessa" jos se on välillä early_start..ensimmäinen vahti
    first_shift_start = shift_info.get('shifts', [(0, 0)])[0][0]
    last_shift_end = shift_info.get('shifts', [(0, 48)])[-1][1]
    
    # Aikaisin laajennus (ennen ensimmäistä vahtia)
    if early_start <= slot < first_shift_start or (early_start > first_shift_start and slot >= early_start):
        # Watchman 1: slot 44-47 (22:00-00:00) on aikainen laajennus
        if state['extended_end']:
            return False  # Jo käytetty myöhäinen laajennus
        return True
    
    # Myöhäinen laajennus (jälkeen viimeisen vahdin)
    if last_shift_end <= slot < min(late_end, 48):
        # Esim. Watchman 1: slot 32-35 (16:00-18:00) on myöhäinen laajennus
        if state['extended_start']:
            return False  # Jo käytetty aikainen laajennus
        return True
    
    # Watchman 3 erikoistapaus: late_end=52 menee yli keskiyön
    if late_end > 48 and slot < (late_end - 48):
        if state['extended_start']:
            return False
        return True
    
    # Slotin ulkopuolella watchmanin mahdollista työaikaa
    return False


def update_watchman_state(watchman, slot, watchman_states):
    """
    Päivittää watchmanin tilan kun hänelle annetaan slotti vahtivuoron ulkopuolella.
    """
    if watchman not in watchman_states:
        watchman_states[watchman] = {'extended_start': False, 'extended_end': False}
    
    # Jos slotti on vahtivuoron sisällä, ei päivitetä tilaa
    if is_within_watchman_shift(watchman, slot):
        return
    
    shift_info = WATCHMAN_SHIFTS.get(watchman, {})
    early_start = shift_info.get('early_start', 0)
    first_shift_start = shift_info.get('shifts', [(0, 0)])[0][0]
    last_shift_end = shift_info.get('shifts', [(0, 48)])[-1][1]
    late_end = shift_info.get('late_end', 48)
    
    # Aikaisin laajennus
    if early_start <= slot < first_shift_start or (early_start > first_shift_start and slot >= early_start):
        watchman_states[watchman]['extended_start'] = True
    # Myöhäinen laajennus
    elif last_shift_end <= slot < min(late_end, 48):
        watchman_states[watchman]['extended_end'] = True
    # Watchman 3 yli keskiyön
    elif late_end > 48 and slot < (late_end - 48):
        watchman_states[watchman]['extended_end'] = True


def find_available_watchman(slot, watchman_states, wm_work, max_hours=MAX_HOURS):
    """
    Etsii watchmanin joka voi ottaa slotin.
    
    Returns:
        Watchmanin nimi tai None
    """
    for wm in WATCHMEN:
        if wm_work[wm][slot]:
            continue  # Jo töissä tässä slotissa
        
        current_hours = sum(wm_work[wm]) / 2
        if current_hours >= max_hours:
            continue  # Max tunnit täynnä
        
        if can_watchman_take_slot(wm, slot, watchman_states):
            return wm
    
    return None


def would_cause_stcw_violation(dm, new_slots, dm_work, prev_day_work, min_longest_rest_hours=6):
    """
    Tarkistaa aiheuttaisiko uusien slottien lisääminen STCW-rikkeen.
    
    Tarkistaa sekä edellinen->nykyinen että nykyinen->seuraava päivä.
    Huomioi myös carry-overin (slotit jotka menevät yli keskiyön).
    
    Args:
        dm: Dayman
        new_slots: Lista sloteista jotka lisättäisiin
        dm_work: Nykyiset työvuorot (dict tai lista)
        prev_day_work: Edellisen päivän työvuorot (dict tai lista)
        
    Returns:
        True jos aiheuttaisi rikkeen
    """
    # Hae nykyinen työvuoro
    if isinstance(dm_work, dict):
        current_work = dm_work[dm][:]
    else:
        current_work = dm_work[:]
    
    # Laske carry-over slotit (yli keskiyön menevät)
    carry_over_slots = []
    for slot in new_slots:
        if slot >= 48:
            carry_over_slots.append(slot - 48)
        elif 0 <= slot < 48:
            current_work[slot] = True
    
    # Hae edellisen päivän työvuoro
    if isinstance(prev_day_work, dict):
        prev_work = prev_day_work.get(dm, [False] * 48)
    else:
        prev_work = prev_day_work if prev_day_work else [False] * 48
    
    # Tarkista edellinen->nykyinen
    ok1, _, _ = check_stcw_sliding(prev_work, current_work, min_longest_rest_hours)
    if not ok1:
        return True
    
    # Tarkista nykyinen->seuraava (oletus: normaali 08-16 työpäivä + carry-over)
    next_day_work = [False] * 48
    
    # Lisää carry-over slotit
    for slot in carry_over_slots:
        if 0 <= slot < 48:
            next_day_work[slot] = True
    
    # Lisää normaali työpäivä (08-16, lounas pois)
    for slot in range(NORMAL_START, NORMAL_END):
        if LUNCH_START <= slot < LUNCH_END:
            continue
        next_day_work[slot] = True
    
    ok2, _, _ = check_stcw_sliding(current_work, next_day_work, min_longest_rest_hours)
    if not ok2:
        return True
    
    return False


# ============================================================================
# STCW-TARKISTUS
# ============================================================================

def analyze_stcw_window(window_48_slots):
    """
    Analysoi STCW-lepoajat yhdestä 24h (48 slotin) ikkunasta.
    """
    rest_slots = [not w for w in window_48_slots]
    
    rest_periods = []
    current_rest = 0
    for is_rest in rest_slots:
        if is_rest:
            current_rest += 1
        else:
            if current_rest >= 2:
                rest_periods.append(current_rest / 2)
            current_rest = 0
    if current_rest >= 2:
        rest_periods.append(current_rest / 2)
    
    # Yhdistä lepo jos se jatkuu ikkunan alusta loppuun (wrap-around)
    if len(rest_periods) >= 2 and len(window_48_slots) == 48:
        if rest_slots[0] and rest_slots[-1]:
            # Ensimmäinen ja viimeinen slotti ovat lepoa -> yhdistä
            combined = rest_periods[-1] + rest_periods[0]
            rest_periods = [combined] + rest_periods[1:-1]
    
    total_rest = sum(rest_periods)
    longest_rest = max(rest_periods) if rest_periods else 0
    
    return {
        'total_rest': total_rest,
        'longest_rest': longest_rest,
        'rest_periods': rest_periods
    }


def analyze_stcw(work_48h):
    """
    Analysoi STCW-lepoajat 48h (2 päivän) työvuorolistasta.
    Käyttää vain ensimmäistä 24h ikkunaa (taaksepäin yhteensopivuus).
    """
    window = work_48h[:48] if len(work_48h) >= 48 else work_48h
    analysis = analyze_stcw_window(window)
    
    # Lisää top_two taaksepäin yhteensopivuutta varten
    sorted_periods = sorted(analysis['rest_periods'], reverse=True)
    top_two = sorted_periods[:2] if len(sorted_periods) >= 2 else sorted_periods
    analysis['top_two_total'] = sum(top_two)
    
    return analysis


def check_stcw_sliding(prev_day_work, current_day_work, min_longest_rest_hours=6):
    """
    Tarkistaa STCW-vaatimukset liukuvalla 24h ikkunalla.
    
    Käy läpi kaikki 48 mahdollista 24h jaksoa jotka alkavat
    edellisen päivän jokaisesta puolen tunnin slotista.
    
    Returns:
        (ok, worst_slot, worst_analysis)
        - ok: True jos kaikki ikkunat OK
        - worst_slot: Slotin indeksi jossa pahin rike (tai None)
        - worst_analysis: Analyysi pahimmasta ikkunasta (tai None)
    """
    if prev_day_work is None:
        prev_day_work = [False] * 48
    if current_day_work is None:
        current_day_work = [False] * 48
    
    combined = prev_day_work + current_day_work  # 96 slottia
    
    worst_slot = None
    worst_total_rest = 999
    worst_analysis = None
    
    # Tarkista jokainen 24h ikkuna (48 ikkunaa)
    for start in range(48):
        window = combined[start:start + 48]
        analysis = analyze_stcw_window(window)
        
        # Tarkista rike
        if analysis['total_rest'] < 10 or analysis['longest_rest'] < min_longest_rest_hours:
            # Tallenna pahin tapaus
            if analysis['total_rest'] < worst_total_rest:
                worst_total_rest = analysis['total_rest']
                worst_slot = start
                worst_analysis = analysis
    
    if worst_slot is not None:
        return False, worst_slot, worst_analysis
    
    return True, None, None


def check_stcw_ok(work_slots, prev_day_work=None, min_longest_rest_hours=6):
    """
    Tarkistaa täyttääkö työvuoro STCW-vaatimukset.
    Käyttää liukuvaa 24h ikkunaa.
    """
    if prev_day_work is None:
        prev_day_work = [False] * 48
    
    ok, _, _ = check_stcw_sliding(prev_day_work, work_slots, min_longest_rest_hours)
    return ok


def check_stcw_at_slot(work_48h, end_slot, min_longest_rest_hours=6):
    """
    Tarkistaa STCW-statuksen tietyssä kohdassa.
    """
    start_slot = max(0, end_slot - 47)
    window = work_48h[start_slot:end_slot + 1]
    
    while len(window) < 48:
        window = [False] + window
    
    analysis = analyze_stcw(window)
    
    total_ok = analysis['total_rest'] >= 10
    longest_ok = analysis['longest_rest'] >= min_longest_rest_hours
    
    if total_ok and longest_ok:
        status = "OK"
    elif not total_ok:
        status = "TOTAL_REST_VIOLATION"
    else:
        status = "LONGEST_REST_VIOLATION"
    
    return {
        'status': status,
        'total_rest': analysis['total_rest'],
        'longest_rest': analysis['longest_rest'],
        'rest_periods': analysis['rest_periods']
    }


def find_earliest_valid_start(prev_day_work, min_hours=MIN_HOURS, min_longest_rest_hours=6):
    """
    Etsii aikaisimman aloitusslottin jolla työpäivä ei riko STCW:tä.
    
    Simuloi työpäivän eri aloitusajoilla ja tarkistaa STCW:n.
    Olettaa että työntekijä tekee min_hours tunnin työpäivän.
    
    Args:
        prev_day_work: Edellisen päivän työvuoro (48 slottia)
        min_hours: Tavoitetunnit (oletus 8h)
        min_longest_rest_hours: STCW pisin lepo vaatimus
    
    Returns:
        Aikaisin mahdollinen aloitusslotti (esim. 16 = 08:00, 18 = 09:00)
    """
    if prev_day_work is None:
        return NORMAL_START
    
    # Jos edellispäivänä ei töitä, voi aloittaa normaalisti
    if not any(prev_day_work):
        return NORMAL_START
    
    target_slots = int(min_hours * 2)  # 8h = 16 slottia
    
    # Kokeile eri aloitusaikoja
    for start_slot in range(NORMAL_START, NORMAL_END):
        # Simuloi työpäivä tästä aloituksesta
        test_work = [False] * 48
        slots_added = 0
        slot = start_slot
        
        while slots_added < target_slots and slot < 48:
            # Ohita lounas
            if LUNCH_START <= slot < LUNCH_END:
                slot += 1
                continue
            test_work[slot] = True
            slots_added += 1
            slot += 1
        
        # Tarkista STCW
        ok, _, _ = check_stcw_sliding(prev_day_work, test_work, min_longest_rest_hours)
        
        if ok:
            return start_slot
    
    # Jos mikään aloitusaika ei toimi normaalityöaikana, 
    # palauta NORMAL_START ja anna fix_stcw_violations hoitaa
    return NORMAL_START


# ============================================================================
# RAJOITTEIDEN KÄSITTELY
# ============================================================================

def can_work_slot(worker, slot, day_idx, constraints, current_hours=0):
    """Tarkistaa voiko työntekijä tehdä tietyn slotin."""
    if not constraints:
        return True
    
    for c in constraints:
        c_worker = c.get("worker")
        c_type = c.get("type")
        c_day = c.get("day")
        
        if c_worker and c_worker != worker:
            continue
        if c_day is not None and c_day != day_idx + 1:
            continue
        
        if c_type == "no_night_shift":
            if slot < NORMAL_START:
                return False
        elif c_type == "no_evening_shift":
            if slot >= NORMAL_END:
                return False
        elif c_type == "cannot_work_slot":
            start = parse_time_str(c.get("start_time"))
            end = parse_time_str(c.get("end_time"))
            if start is not None and end is not None:
                if start <= slot < end:
                    return False
        elif c_type == "max_hours":
            max_h = c.get("value", MAX_HOURS)
            if current_hours >= max_h:
                return False
        elif c_type == "day_off":
            return False
    
    return True


def must_work_slot(worker, slot, day_idx, constraints):
    """Tarkistaa onko työntekijän pakko tehdä tietty slotti."""
    if not constraints:
        return False
    
    for c in constraints:
        c_worker = c.get("worker")
        c_type = c.get("type")
        c_day = c.get("day")
        
        if c_worker and c_worker != worker:
            continue
        if c_day is not None and c_day != day_idx + 1:
            continue
        
        if c_type == "must_work_slot":
            start = parse_time_str(c.get("start_time"))
            end = parse_time_str(c.get("end_time"))
            if start is not None and end is not None:
                if start <= slot < end:
                    return True
    
    return False


def is_day_off(worker, day_idx, constraints):
    """Tarkistaa onko työntekijällä vapaapäivä."""
    if not constraints:
        return False
    
    for c in constraints:
        if c.get("type") == "day_off":
            if c.get("worker") == worker:
                c_day = c.get("day")
                if c_day is None or c_day == day_idx + 1:
                    return True
    return False


def get_min_hours(worker, constraints):
    """Palauttaa työntekijän minimitunnit."""
    for c in constraints or []:
        if c.get("worker") == worker and c.get("type") == "min_hours":
            return c.get("value", MIN_HOURS)
    return MIN_HOURS


def get_max_hours(worker, constraints):
    """Palauttaa työntekijän maksimitunnit."""
    for c in constraints or []:
        if c.get("worker") == worker and c.get("type") == "max_hours":
            return c.get("value", MAX_HOURS)
    return MAX_HOURS


def get_preferred_night_worker(day_idx, constraints, daymen):
    """Palauttaa yövuoroon määrätyn työntekijän."""
    if not constraints:
        return None
    
    for c in constraints:
        if c.get("type") == "assign_night_shift":
            c_day = c.get("day")
            if c_day is None or c_day == day_idx + 1:
                worker = c.get("worker")
                if worker in daymen:
                    return worker
    return None


# ============================================================================
# VAIHE 0: JATKUVIEN YÖVUOROJEN ANALYYSI
# ============================================================================

def analyze_continuous_nights(days_data):
    """
    Analysoi jatkuvat yövuorot etukäteen.
    Palauttaa listan jatkuvista öistä.
    """
    continuous_nights = []
    num_days = len(days_data)
    
    for d in range(num_days - 1):
        curr = days_data[d]
        next_day = days_data[d + 1]
        
        curr_port_ops = curr.get('port_operations') or []
        next_port_ops = next_day.get('port_operations') or []
        curr_op_end = curr_port_ops[-1].get('end_hour') if curr_port_ops else curr.get('port_op_end_hour')
        next_op_start = next_port_ops[0].get('start_hour') if next_port_ops else next_day.get('port_op_start_hour')
        
        if curr_op_end == 0 and next_op_start == 0:
            continuous_nights.append({
                'day_index': d,
            })
    
    return continuous_nights


def evaluate_night_split(prev_early, prev_late, split_slot, arrival_start=None, departure_start=None, min_longest_rest_hours=6):
    """
    Arvioi yövuoron jakokohdan hyvyyttä.
    """
    early_night = list(range(0, split_slot))
    late_night = list(range(split_slot, NORMAL_START))
    
    early_combined = prev_early + [True if i in early_night else False for i in range(48)]
    late_combined = prev_late + [True if i in late_night else False for i in range(48)]
    
    early_analysis = analyze_stcw(early_combined)
    late_analysis = analyze_stcw(late_combined)
    
    total_issues = 0
    if early_analysis['total_rest'] < 10:
        total_issues += 1
    if early_analysis['longest_rest'] < min_longest_rest_hours:
        total_issues += 1
    if late_analysis['total_rest'] < 10:
        total_issues += 1
    if late_analysis['longest_rest'] < min_longest_rest_hours:
        total_issues += 1
    
    min_longest_rest = min(early_analysis['longest_rest'], late_analysis['longest_rest'])
    min_total_rest = min(early_analysis['total_rest'], late_analysis['total_rest'])
    
    return (total_issues, -min_longest_rest, -min_total_rest)


def choose_night_split_slot(prev_early, prev_late, arrival_start=None, departure_start=None, min_longest_rest_hours=6):
    """
    Valitsee optimaalisen yövuoron jakokohdan (01:00 - 07:00 väliltä).
    """
    candidate_slots = list(range(time_to_slot(1, 0), time_to_slot(7, 0) + 1))
    best_slot = time_to_slot(3, 0)
    best_score = None
    
    for split_slot in candidate_slots:
        score = evaluate_night_split(
            prev_early,
            prev_late,
            split_slot,
            arrival_start,
            departure_start,
            min_longest_rest_hours=min_longest_rest_hours
        )
        if best_score is None or score < best_score:
            best_score = score
            best_slot = split_slot
    
    return best_slot


# ============================================================================
# VAIHE 1: PAKOLLISET SLOTIT
# ============================================================================

def parse_day_times(info):
    """
    Parsii päivän ajat sloteiksi.
    Palauttaa dictin kaikista relevanteista sloteista.
    """
    port_operations_info = info.get('port_operations')
    if port_operations_info:
        port_operations = []
        for operation in port_operations_info:
            op_start_h = operation.get('start_hour')
            op_start_m = operation.get('start_minute', 0)
            op_end_h = operation.get('end_hour')
            op_end_m = operation.get('end_minute', 0)
            if op_start_h is None:
                continue
            op_start = time_to_slot(op_start_h, op_start_m)
            if op_end_h is not None and op_end_h < op_start_h:
                op_end = 48
            elif op_end_h == 0 and op_start_h > 0:
                op_end = 48
            elif op_end_h is not None:
                op_end = time_to_slot(op_end_h, op_end_m)
            else:
                op_end = NORMAL_END
            port_operations.append((op_start, op_end))
    else:
        op_start_h = info.get('port_op_start_hour')
        op_start_m = info.get('port_op_start_minute', 0)
        op_end_h = info.get('port_op_end_hour')
        op_end_m = info.get('port_op_end_minute', 0)
        if op_start_h is not None:
            op_start = time_to_slot(op_start_h, op_start_m)
            if op_end_h is not None and op_end_h < op_start_h:
                op_end = 48
            elif op_end_h == 0 and op_start_h > 0:
                op_end = 48
            elif op_end_h is not None:
                op_end = time_to_slot(op_end_h, op_end_m)
            else:
                op_end = NORMAL_END
            port_operations = [(op_start, op_end)]
        else:
            # Ei operaatioita määritelty -> tyhjä lista
            port_operations = []
            op_start = None
            op_end = None

    # Tulo: syötetty aika on LOPPUMISAIKA, operaatio kestää 1h (2 slottia)
    # Eli aloitus = syötetty aika - 2 slottia
    arrivals = info.get('arrivals')
    if arrivals:
        arrival_starts = []
        for entry in arrivals:
            if entry.get('hour') is not None:
                end_slot = time_to_slot(entry['hour'], entry.get('minute', 0))
                start_slot = max(0, end_slot - 2)  # 1h = 2 slottia ennen
                arrival_starts.append(start_slot)
    else:
        arrival_h = info.get('arrival_hour')
        arrival_m = info.get('arrival_minute', 0)
        if arrival_h is not None:
            end_slot = time_to_slot(arrival_h, arrival_m)
            start_slot = max(0, end_slot - 2)  # 1h = 2 slottia ennen
            arrival_starts = [start_slot]
        else:
            arrival_starts = []

    departures = info.get('departures')
    if departures:
        departure_starts = [time_to_slot(entry['hour'], entry.get('minute', 0)) for entry in departures if entry.get('hour') is not None]
    else:
        departure_h = info.get('departure_hour')
        departure_m = info.get('departure_minute', 0)
        departure_starts = [time_to_slot(departure_h, departure_m)] if departure_h is not None else []

    # Slussi tulo: syötetty aika on LOPPUMISAIKA, operaatio kestää 2.5h (5 slottia)
    # Eli aloitus = syötetty aika - 5 slottia
    sluice_arrivals = info.get('sluice_arrivals')
    if sluice_arrivals:
        sluice_arr_starts = []
        for entry in sluice_arrivals:
            if entry.get('hour') is not None:
                end_slot = time_to_slot(entry['hour'], entry.get('minute', 0))
                start_slot = max(0, end_slot - 5)  # 2.5h = 5 slottia ennen
                sluice_arr_starts.append(start_slot)
    else:
        sluice_arr_h = info.get('sluice_arrival_hour')
        sluice_arr_m = info.get('sluice_arrival_minute', 0)
        if sluice_arr_h is not None:
            end_slot = time_to_slot(sluice_arr_h, sluice_arr_m)
            start_slot = max(0, end_slot - 5)  # 2.5h = 5 slottia ennen
            sluice_arr_starts = [start_slot]
        else:
            sluice_arr_starts = []

    # Lähtöslussi pysyy ennallaan - syötetty aika on aloitusaika
    sluice_departures = info.get('sluice_departures')
    if sluice_departures:
        sluice_dep_starts = [time_to_slot(entry['hour'], entry.get('minute', 0)) for entry in sluice_departures if entry.get('hour') is not None]
    else:
        sluice_dep_h = info.get('sluice_departure_hour')
        sluice_dep_m = info.get('sluice_departure_minute', 0)
        sluice_dep_starts = [time_to_slot(sluice_dep_h, sluice_dep_m)] if sluice_dep_h is not None else []

    shiftings = info.get('shiftings')
    if shiftings:
        shifting_starts = [time_to_slot(entry['hour'], entry.get('minute', 0)) for entry in shiftings if entry.get('hour') is not None]
    else:
        shifting_h = info.get('shifting_hour')
        shifting_m = info.get('shifting_minute', 0)
        shifting_starts = [time_to_slot(shifting_h, shifting_m)] if shifting_h is not None else []

    # Laske op_start ja op_end jos operaatioita on
    if port_operations:
        op_start = min(start for start, _ in port_operations)
        op_end = max(end for _, end in port_operations)
    else:
        op_start = None
        op_end = None

    return {
        'op_start': op_start,
        'op_end': op_end,
        'op_ranges': port_operations,
        'arrival_start': arrival_starts[0] if arrival_starts else None,
        'arrival_starts': arrival_starts,
        'departure_start': departure_starts[0] if departure_starts else None,
        'departure_starts': departure_starts,
        'sluice_arr_start': sluice_arr_starts[0] if sluice_arr_starts else None,
        'sluice_arr_starts': sluice_arr_starts,
        'sluice_dep_start': sluice_dep_starts[0] if sluice_dep_starts else None,
        'sluice_dep_starts': sluice_dep_starts,
        'shifting_start': shifting_starts[0] if shifting_starts else None,
        'shifting_starts': shifting_starts,
    }


def is_op_slot(times, slot):
    """Palauttaa True jos slot kuuluu johonkin satamaoperaatiojaksoon."""
    return any(start <= slot < min(end, 48) for start, end in times.get('op_ranges', []))


def apply_constraint_slots(dm_work, dm_ops, daymen, day_idx, times, constraints):
    """
    Vaihe 0.5: Lisää pakolliset slotit rajoitteista (must_work_slot).
    """
    for dm in daymen:
        if is_day_off(dm, day_idx, constraints):
            continue
        for slot in range(48):
            if must_work_slot(dm, slot, day_idx, constraints):
                dm_work[dm][slot] = True
                if is_op_slot(times, slot):
                    dm_ops[dm][slot] = True


def apply_arrival_slots(dm_work, dm_arr, active_daymen, day_idx, times, constraints):
    """
    Vaihe 1.1: Tulo - kaikki aktiiviset daymanit (1h).
    """
    for arrival_start in times.get('arrival_starts', []):
        for dm in active_daymen:
            can_add = True
            for slot in range(arrival_start, arrival_start + 2):
                if not can_work_slot(dm, slot, day_idx, constraints, sum(dm_work[dm])/2):
                    can_add = False
                    break
            if can_add:
                add_block(dm_work[dm], arrival_start, arrival_start + 2, dm_arr[dm])


def apply_departure_slots(dm_work, dm_dep, active_daymen, day_idx, times, constraints):
    """
    Vaihe 1.2: Lähtö - 2 daymaniä (1h).
    """
    for departure_start in times.get('departure_starts', []):
        scores = {}
        for dm in active_daymen:
            can_do = True
            for slot in range(departure_start, departure_start + 2):
                if not can_work_slot(dm, slot, day_idx, constraints, sum(dm_work[dm])/2):
                    can_do = False
                    break
            if not can_do:
                continue

            hours = sum(dm_work[dm]) / 2
            continuity = 1 if (departure_start > 0 and dm_work[dm][departure_start - 1]) else 0
            scores[dm] = -hours + continuity

        selected = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)[:2]
        for dm in selected:
            add_block(dm_work[dm], departure_start, departure_start + 2, dm_dep[dm])


def apply_sluice_arrival_slots(dm_work, dm_sluice, daymen, times, pending_next_day=None,
                                prev_day_work=None, wm_work=None, wm_sluice=None, 
                                watchman_states=None, min_longest_rest_hours=6):
    """
    Vaihe 1.3: Slussi tulo - 1. tunti 2 henkilöä, loput 1.5h 3 henkilöä (2.5h kokonaan).
    
    Käyttää ensisijaisesti daymaneita, mutta jos STCW-rike uhkaa, käyttää watchmaneja.
    
    Args:
        pending_next_day: Dict johon tallennetaan ylivuoto seuraavalle päivälle
        prev_day_work: Edellisen päivän työvuorot (STCW-tarkistukseen)
        wm_work: Watchmanien työvuorot
        wm_sluice: Watchmanien sluice-merkinnät
        watchman_states: Watchmanien extended_start/end tilat
    
    Returns:
        pending_next_day päivitettynä
    """
    if pending_next_day is None:
        pending_next_day = {dm: {'work': [], 'sluice': []} for dm in daymen}
    if prev_day_work is None:
        prev_day_work = {dm: [False] * 48 for dm in daymen}
    if watchman_states is None:
        watchman_states = {wm: {'extended_start': False, 'extended_end': False} for wm in WATCHMEN}
    
    for sluice_arr_start in times.get('sluice_arr_starts', []):
        # Kaikki slussin slotit (myös yli keskiyön menevät) STCW-tarkistusta varten
        sluice_slots_full = list(range(sluice_arr_start, sluice_arr_start + 5))
        
        # Pisteytetään daymanit
        scores = {}
        for dm in daymen:
            hours = sum(dm_work[dm]) / 2
            # Tarkista aiheuttaisiko STCW-rikkeen (käytä kaikkia slotteja)
            would_violate = would_cause_stcw_violation(
                dm, sluice_slots_full, dm_work, prev_day_work, min_longest_rest_hours
            )
            # Rankaistaan niitä jotka aiheuttaisivat rikkeen
            stcw_penalty = -1000 if would_violate else 0
            scores[dm] = -hours + stcw_penalty

        sorted_daymen = sorted(daymen, key=lambda x: scores[x], reverse=True)
        
        # 1. tunti (2 slottia) - valitse 2 henkilöä
        first_hour_workers = []
        for dm in sorted_daymen:
            if len(first_hour_workers) >= 2:
                break
            # Tarkista STCW
            if not would_cause_stcw_violation(dm, [sluice_arr_start, sluice_arr_start + 1], 
                                               dm_work, prev_day_work, min_longest_rest_hours):
                first_hour_workers.append(('dayman', dm))
        
        # Jos ei tarpeeksi daymaneita, käytä watchmaneja
        if len(first_hour_workers) < 2 and wm_work is not None:
            for slot in [sluice_arr_start, sluice_arr_start + 1]:
                if slot < 48:
                    wm = find_available_watchman(slot, watchman_states, wm_work)
                    if wm and ('watchman', wm) not in first_hour_workers:
                        first_hour_workers.append(('watchman', wm))
                        if len(first_hour_workers) >= 2:
                            break
        
        # Lisää slotit valituille (1. tunti)
        for worker_type, worker in first_hour_workers[:2]:
            if worker_type == 'dayman':
                overflow = add_block(dm_work[worker], sluice_arr_start, sluice_arr_start + 2, dm_sluice[worker])
                if overflow > 0:
                    for i in range(overflow):
                        pending_next_day[worker]['work'].append(i)
                        pending_next_day[worker]['sluice'].append(i)
            else:  # watchman
                if wm_work is not None and wm_sluice is not None:
                    add_block(wm_work[worker], sluice_arr_start, sluice_arr_start + 2, wm_sluice[worker])
                    for slot in [sluice_arr_start, sluice_arr_start + 1]:
                        if slot < 48:
                            update_watchman_state(worker, slot, watchman_states)

        # Loput 1.5h (3 slottia) - valitse 3 henkilöä
        second_part_workers = []
        for dm in sorted_daymen:
            if len(second_part_workers) >= 3:
                break
            # Käytä kaikkia slotteja (myös yli keskiyön) STCW-tarkistuksessa
            slots_2nd_full = list(range(sluice_arr_start + 2, sluice_arr_start + 5))
            if not would_cause_stcw_violation(dm, slots_2nd_full, dm_work, prev_day_work, min_longest_rest_hours):
                second_part_workers.append(('dayman', dm))
        
        # Jos ei tarpeeksi daymaneita, käytä watchmaneja
        if len(second_part_workers) < 3 and wm_work is not None:
            for slot in range(sluice_arr_start + 2, min(sluice_arr_start + 5, 48)):
                wm = find_available_watchman(slot, watchman_states, wm_work)
                if wm and ('watchman', wm) not in second_part_workers:
                    second_part_workers.append(('watchman', wm))
                    if len(second_part_workers) >= 3:
                        break
        
        # Lisää slotit valituille (loput 1.5h)
        for worker_type, worker in second_part_workers[:3]:
            if worker_type == 'dayman':
                overflow = add_block(dm_work[worker], sluice_arr_start + 2, sluice_arr_start + 5, dm_sluice[worker])
                if overflow > 0:
                    first_overflow_slot = sluice_arr_start + 5 - overflow
                    for i in range(overflow):
                        next_day_slot = (first_overflow_slot + i) - 48
                        if next_day_slot not in pending_next_day[worker]['work']:
                            pending_next_day[worker]['work'].append(next_day_slot)
                            pending_next_day[worker]['sluice'].append(next_day_slot)
            else:  # watchman
                if wm_work is not None and wm_sluice is not None:
                    add_block(wm_work[worker], sluice_arr_start + 2, sluice_arr_start + 5, wm_sluice[worker])
                    for slot in range(sluice_arr_start + 2, min(sluice_arr_start + 5, 48)):
                        update_watchman_state(worker, slot, watchman_states)
    
    return pending_next_day


def apply_sluice_departure_slots(dm_work, dm_sluice, daymen, times, pending_next_day=None,
                                  prev_day_work=None, wm_work=None, wm_sluice=None,
                                  watchman_states=None, min_longest_rest_hours=6):
    """
    Vaihe 1.4: Slussi lähtö - 2 henkilöä (2.5h).
    
    Käyttää ensisijaisesti daymaneita, mutta jos STCW-rike uhkaa, käyttää watchmaneja.
    
    Args:
        pending_next_day: Dict johon tallennetaan ylivuoto seuraavalle päivälle
        prev_day_work: Edellisen päivän työvuorot (STCW-tarkistukseen)
        wm_work: Watchmanien työvuorot
        wm_sluice: Watchmanien sluice-merkinnät
        watchman_states: Watchmanien extended_start/end tilat
    
    Returns:
        pending_next_day päivitettynä
    """
    if pending_next_day is None:
        pending_next_day = {dm: {'work': [], 'sluice': []} for dm in daymen}
    if prev_day_work is None:
        prev_day_work = {dm: [False] * 48 for dm in daymen}
    if watchman_states is None:
        watchman_states = {wm: {'extended_start': False, 'extended_end': False} for wm in WATCHMEN}
    
    for sluice_dep_start in times.get('sluice_dep_starts', []):
        # Kaikki slussin slotit (myös yli keskiyön menevät) STCW-tarkistusta varten
        sluice_slots_full = list(range(sluice_dep_start, sluice_dep_start + 5))
        
        # Pisteytetään daymanit
        scores = {}
        for dm in daymen:
            hours = sum(dm_work[dm]) / 2
            continuity = 1 if (sluice_dep_start > 0 and dm_work[dm][sluice_dep_start - 1]) else 0
            
            # Tarkista onko jo osallistunut slussiin tänään (rankaistaan)
            already_in_sluice = any(dm_sluice[dm])
            sluice_penalty = -500 if already_in_sluice else 0
            
            # Tarkista aiheuttaisiko STCW-rikkeen (käytä kaikkia slotteja)
            would_violate = would_cause_stcw_violation(
                dm, sluice_slots_full, dm_work, prev_day_work, min_longest_rest_hours
            )
            stcw_penalty = -1000 if would_violate else 0
            scores[dm] = -hours + continuity + stcw_penalty + sluice_penalty

        sorted_daymen = sorted(daymen, key=lambda x: scores[x], reverse=True)
        
        # Valitse 2 henkilöä
        selected_workers = []
        for dm in sorted_daymen:
            if len(selected_workers) >= 2:
                break
            if not would_cause_stcw_violation(dm, sluice_slots_full, dm_work, prev_day_work, min_longest_rest_hours):
                selected_workers.append(('dayman', dm))
        
        # Jos ei tarpeeksi daymaneita, käytä watchmaneja
        if len(selected_workers) < 2 and wm_work is not None:
            for slot in range(sluice_dep_start, min(sluice_dep_start + 5, 48)):
                wm = find_available_watchman(slot, watchman_states, wm_work)
                if wm and ('watchman', wm) not in selected_workers:
                    selected_workers.append(('watchman', wm))
                    if len(selected_workers) >= 2:
                        break
        
        # Lisää slotit valituille
        for worker_type, worker in selected_workers[:2]:
            if worker_type == 'dayman':
                overflow = add_block(dm_work[worker], sluice_dep_start, sluice_dep_start + 5, dm_sluice[worker])
                if overflow > 0:
                    first_overflow_slot = sluice_dep_start + 5 - overflow
                    for i in range(overflow):
                        next_day_slot = (first_overflow_slot + i) - 48
                        if next_day_slot not in pending_next_day[worker]['work']:
                            pending_next_day[worker]['work'].append(next_day_slot)
                            pending_next_day[worker]['sluice'].append(next_day_slot)
            else:  # watchman
                if wm_work is not None and wm_sluice is not None:
                    add_block(wm_work[worker], sluice_dep_start, sluice_dep_start + 5, wm_sluice[worker])
                    for slot in range(sluice_dep_start, min(sluice_dep_start + 5, 48)):
                        update_watchman_state(worker, slot, watchman_states)
    
    return pending_next_day


def apply_shifting_slots(dm_work, dm_shifting, daymen, times):
    """
    Vaihe 1.5: Shiftaus - kaikki daymanit (1h).
    """
    for shifting_start in times.get('shifting_starts', []):
        for dm in daymen:
            add_block(dm_work[dm], shifting_start, shifting_start + 2, dm_shifting[dm])


def apply_op_outside_normal_hours(dm_work, dm_ops, active_daymen, day_idx, times, 
                                   constraints, prev_day_work, continuous_night_info,
                                   min_longest_rest_hours=6):
    """
    Vaihe 1.6: Satamaop 08-17 ULKOPUOLELLA - aina 1 dayman töissä.
    """
    op_start = times['op_start']
    op_end = times['op_end']
    
    op_outside_slots = []
    for start, end in times.get('op_ranges', []):
        for slot in range(start, min(end, 48)):
            if (slot < NORMAL_START or slot >= NORMAL_END) and slot != LUNCH_START:
                op_outside_slots.append(slot)
    op_outside_slots = sorted(set(op_outside_slots))
    
    # Jos jatkuva yö edellisestä päivästä
    if continuous_night_info is not None:
        early_worker = continuous_night_info['early_worker']
        late_worker = continuous_night_info['late_worker']
        night_split_slot = continuous_night_info['split_slot']
        
        for slot in range(0, night_split_slot):
            if slot in op_outside_slots:
                dm_work[early_worker][slot] = True
                dm_ops[early_worker][slot] = True
        
        for slot in range(night_split_slot, NORMAL_START):
            if slot in op_outside_slots:
                dm_work[late_worker][slot] = True
                dm_ops[late_worker][slot] = True
        
        op_outside_slots = [s for s in op_outside_slots if s >= NORMAL_START]
    
    current_worker = None
    preferred = get_preferred_night_worker(day_idx, constraints, active_daymen)
    
    for slot in op_outside_slots:
        can_continue = False
        
        if current_worker is not None:
            current_hours = sum(dm_work[current_worker]) / 2
            max_h = get_max_hours(current_worker, constraints)
            
            if current_hours < max_h and can_work_slot(current_worker, slot, day_idx, constraints, current_hours):
                test_work = dm_work[current_worker][:]
                test_work[slot] = True
                stcw_ok = check_stcw_ok(
                    test_work,
                    prev_day_work[current_worker],
                    min_longest_rest_hours=min_longest_rest_hours
                )
                
                if stcw_ok:
                    can_continue = True
        
        if can_continue:
            dm_work[current_worker][slot] = True
            dm_ops[current_worker][slot] = True
        else:
            best_dm = _find_best_worker_for_slot(
                slot, active_daymen, current_worker, dm_work, prev_day_work,
                day_idx, constraints, preferred,
                min_longest_rest_hours=min_longest_rest_hours
            )
            
            if best_dm:
                dm_work[best_dm][slot] = True
                dm_ops[best_dm][slot] = True
                current_worker = best_dm
            elif current_worker is not None:
                current_hours = sum(dm_work[current_worker]) / 2
                max_h = get_max_hours(current_worker, constraints)
                if current_hours < max_h:
                    dm_work[current_worker][slot] = True
                    dm_ops[current_worker][slot] = True


def _find_best_worker_for_slot(slot, active_daymen, current_worker, dm_work, 
                                prev_day_work, day_idx, constraints, preferred=None,
                                min_longest_rest_hours=6):
    """
    Apufunktio: Etsii parhaan työntekijän tietylle slotille.
    """
    best_dm = None
    best_score = -9999
    
    for dm in active_daymen:
        if dm == current_worker:
            continue
        
        current_hours = sum(dm_work[dm]) / 2
        max_h = get_max_hours(dm, constraints)
        
        if current_hours >= max_h:
            continue
        
        if not can_work_slot(dm, slot, day_idx, constraints, current_hours):
            continue
        
        test_work = dm_work[dm][:]
        test_work[slot] = True
        stcw_ok = check_stcw_ok(
            test_work,
            prev_day_work[dm],
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        if not stcw_ok:
            continue
        
        score = 0
        score += (max_h - current_hours) * 10
        
        if preferred and dm == preferred:
            score += 500
        
        prev_work = prev_day_work[dm]
        last_work_slot = -1
        for s in range(47, -1, -1):
            if prev_work[s]:
                last_work_slot = s
                break
        
        if last_work_slot >= 0:
            rest_slots = (slot + 48) - last_work_slot
            rest_hours = rest_slots / 2
            score += min(rest_hours, 24) * 5
        else:
            score += 24 * 5
        
        if score > best_score:
            best_score = score
            best_dm = dm
    
    return best_dm


# ============================================================================
# VAIHE 3: TYÖBLOKKIEN JAKAMINEN
# ============================================================================

def fill_op_inside_normal_hours(dm_work, dm_ops, active_daymen, day_idx, times, 
                                 constraints, prev_day_work, min_longest_rest_hours=6):
    """
    Vaihe 3.1: Op-kattavuus 08-17 välillä.
    """
    op_start = times['op_start']
    op_end = times['op_end']
    
    op_inside_slots = []
    for start, end in times.get('op_ranges', []):
        for slot in range(max(start, NORMAL_START), min(end, NORMAL_END)):
            if LUNCH_START <= slot < LUNCH_END:
                continue
            op_inside_slots.append(slot)
    op_inside_slots = sorted(set(op_inside_slots))
    
    for slot in op_inside_slots:
        workers_in_slot = [dm for dm in active_daymen if dm_work[dm][slot]]
        
        if len(workers_in_slot) >= 1:
            for dm in workers_in_slot:
                dm_ops[dm][slot] = True
            continue
        
        best_dm = _find_best_worker_for_inside_slot(
            slot, active_daymen, dm_work, prev_day_work, day_idx, constraints,
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        if best_dm:
            dm_work[best_dm][slot] = True
            dm_ops[best_dm][slot] = True
    
    return op_inside_slots


def _find_best_worker_for_inside_slot(slot, active_daymen, dm_work, prev_day_work, 
                                       day_idx, constraints, min_longest_rest_hours=6):
    """
    Apufunktio: Etsii parhaan työntekijän 08-17 slotille.
    Huomioi STCW:n mukaisen aikaisimman aloitusajan.
    """
    best_dm = None
    best_score = -9999
    
    for dm in active_daymen:
        current_hours = sum(dm_work[dm]) / 2
        max_h = get_max_hours(dm, constraints)
        
        if current_hours >= max_h:
            continue
        
        if not can_work_slot(dm, slot, day_idx, constraints, current_hours):
            continue
        
        # Tarkista aikaisin mahdollinen aloitus STCW:n perusteella
        earliest_start = find_earliest_valid_start(
            prev_day_work.get(dm, [False] * 48),
            min_hours=get_min_hours(dm, constraints),
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        # Jos slotti on ennen aikaisinta sallittua aloitusta, ohita
        if slot < earliest_start:
            continue
        
        score = 0
        
        if slot > 0 and dm_work[dm][slot - 1]:
            score += 200
        if slot < 47 and dm_work[dm][slot + 1]:
            score += 200
        
        test_work = dm_work[dm][:]
        test_work[slot] = True
        stcw_ok = check_stcw_ok(
            test_work,
            prev_day_work.get(dm, [False] * 48),
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        if not stcw_ok:
            continue
        
        score += (max_h - current_hours) * 10
        
        if score > best_score:
            best_score = score
            best_dm = dm
    
    return best_dm


def fill_remaining_hours(dm_work, dm_ops, active_daymen, day_idx, times, constraints,
                         prev_day_work=None, min_longest_rest_hours=6):
    """
    Vaihe 3.2: Täytä loput tunnit aikaisimmasta sallitusta aloitusajasta.
    Huomioi STCW:n mukaisen aikaisimman aloitusajan.
    """
    op_start = times['op_start']
    op_end = times['op_end']
    
    if prev_day_work is None:
        prev_day_work = {dm: [False] * 48 for dm in active_daymen}
    
    for dm in active_daymen:
        current_hours = sum(dm_work[dm]) / 2
        min_h = get_min_hours(dm, constraints)
        max_h = get_max_hours(dm, constraints)
        
        if current_hours >= min_h:
            continue
        
        night_work_slots = sum(1 for s in range(0, NORMAL_START) if dm_work[dm][s])
        did_night_shift = night_work_slots >= 4
        
        if did_night_shift:
            _extend_night_shift(dm, dm_work, dm_ops, op_start, op_end, min_h, max_h, 
                               day_idx, constraints)
            continue
        
        # Laske aikaisin mahdollinen aloitus STCW:n perusteella
        earliest_start = find_earliest_valid_start(
            prev_day_work.get(dm, [False] * 48),
            min_hours=min_h,
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        # Aloita aikaisimmasta sallitusta slotista
        slot = max(NORMAL_START, earliest_start)
        
        while current_hours < min_h and slot < NORMAL_END:
            if LUNCH_START <= slot < LUNCH_END:
                slot += 1
                continue
            
            if dm_work[dm][slot]:
                slot += 1
                continue
            
            if not can_work_slot(dm, slot, day_idx, constraints, current_hours):
                slot += 1
                continue
            
            if current_hours >= max_h:
                break
            
            dm_work[dm][slot] = True
            if is_op_slot(times, slot):
                dm_ops[dm][slot] = True
            current_hours = sum(dm_work[dm]) / 2
            slot += 1
        
        # Jos tunnit eivät riitä, jatka iltapäivään/iltaan (NORMAL_END jälkeen)
        if current_hours < min_h:
            for slot in range(NORMAL_END, 48):
                if current_hours >= min_h:
                    break
                if dm_work[dm][slot]:
                    continue
                if not can_work_slot(dm, slot, day_idx, constraints, current_hours):
                    continue
                
                # Tarkista STCW ennen lisäämistä
                test_work = dm_work[dm][:]
                test_work[slot] = True
                if not check_stcw_ok(test_work, prev_day_work.get(dm, [False] * 48), min_longest_rest_hours):
                    continue
                
                dm_work[dm][slot] = True
                if is_op_slot(times, slot):
                    dm_ops[dm][slot] = True
                current_hours = sum(dm_work[dm]) / 2


def _extend_night_shift(dm, dm_work, dm_ops, op_start, op_end, min_h, max_h, 
                        day_idx, constraints):
    """
    Apufunktio: Laajentaa yövuoroa tarvittaessa.
    """
    current_hours = sum(dm_work[dm]) / 2
    
    last_night_slot = -1
    for s in range(NORMAL_START - 1, -1, -1):
        if dm_work[dm][s]:
            last_night_slot = s
            break
    
    if last_night_slot >= 0 and last_night_slot < NORMAL_START - 1:
        for s in range(last_night_slot + 1, NORMAL_START):
            if current_hours >= min_h:
                break
            if current_hours >= max_h:
                break
            if not can_work_slot(dm, s, day_idx, constraints, current_hours):
                break
            dm_work[dm][s] = True
            if is_op_slot(times, s):
                dm_ops[dm][s] = True
            current_hours = sum(dm_work[dm]) / 2


def ensure_op_coverage(dm_work, dm_ops, op_inside_slots, active_daymen, day_idx, constraints):
    """
    Vaihe 3.3: Varmista op-kattavuus uudelleen.
    """
    for slot in op_inside_slots:
        workers_in_slot = [dm for dm in active_daymen if dm_work[dm][slot]]
        
        if len(workers_in_slot) >= 1:
            for dm in workers_in_slot:
                dm_ops[dm][slot] = True
            continue
        
        best_dm = None
        best_score = -9999
        
        for dm in active_daymen:
            current_hours = sum(dm_work[dm]) / 2
            max_h = get_max_hours(dm, constraints)
            
            if current_hours >= max_h:
                continue
            
            if not can_work_slot(dm, slot, day_idx, constraints, current_hours):
                continue
            
            score = 0
            if slot > 0 and dm_work[dm][slot - 1]:
                score += 200
            if slot < 47 and dm_work[dm][slot + 1]:
                score += 200
            score += (max_h - current_hours) * 10
            
            if score > best_score:
                best_score = score
                best_dm = dm
        
        if best_dm:
            dm_work[best_dm][slot] = True
            dm_ops[best_dm][slot] = True


def fill_gaps_between_blocks(dm_work, dm_ops, active_daymen, day_idx, times, constraints):
    """
    Vaihe 3.4: Täytä turhat aukot blokkien välissä.
    """
    op_start = times['op_start']
    op_end = times['op_end']
    
    for dm in active_daymen:
        work = dm_work[dm]
        blocks = get_work_blocks(work)
        max_h = get_max_hours(dm, constraints)
        
        for i in range(len(blocks) - 1):
            _, block1_end = blocks[i]
            block2_start, _ = blocks[i + 1]
            
            gap = block2_start - block1_end
            
            if 0 < gap <= 4 and block1_end >= NORMAL_START and block2_start <= NORMAL_END:
                current_hours = sum(work) / 2
                
                for s in range(block1_end, block2_start):
                    if LUNCH_START <= s < LUNCH_END:
                        continue
                    if current_hours >= max_h:
                        break
                    if not can_work_slot(dm, s, day_idx, constraints, current_hours):
                        continue
                    work[s] = True
                    if is_op_slot(times, s):
                        dm_ops[dm][s] = True
                    current_hours = sum(work) / 2


# ============================================================================
# VAIHE 4: AUKKOJEN TÄYTTÖ
# ============================================================================

def fill_small_gaps(dm_work, dm_ops, active_daymen, times):
    """
    Täytä pienet aukot (max 2h = 4 slottia).
    
    Käy läpi kaikki aukot työjaksojen välissä ja täyttää ne jos:
    - Aukko on max 4 slottia (2h), EI lasketa lounasta mukaan
    - Lounaslotteja ei täytetä
    """
    for dm in active_daymen:
        work = dm_work[dm]
        
        # Käy läpi kaikki slotit ja etsi aukot
        for slot in range(48):
            if work[slot]:
                continue  # Jo töissä
            
            if LUNCH_START <= slot < LUNCH_END:
                continue  # Lounastauko, älä täytä
            
            # Etsi lähin työ vasemmalla ja oikealla
            left_work = -1
            for s in range(slot - 1, -1, -1):
                if work[s]:
                    left_work = s
                    break
            
            right_work = -1
            for s in range(slot + 1, 48):
                if work[s]:
                    right_work = s
                    break
            
            # Jos työtä molemmilla puolilla
            if left_work >= 0 and right_work >= 0:
                # Laske aukon koko ILMAN lounasslotteja
                gap_slots = []
                for s in range(left_work + 1, right_work):
                    if LUNCH_START <= s < LUNCH_END:
                        continue  # Ohita lounas
                    if not work[s]:
                        gap_slots.append(s)
                
                gap_size = len(gap_slots)
                
                if gap_size <= 4 and slot in gap_slots:
                    # Täytä tämä slotti
                    work[slot] = True
                    if is_op_slot(times, slot):
                        dm_ops[dm][slot] = True


def rebalance_dayman_hours(
    dm_work,
    dm_ops,
    dm_arr,
    dm_dep,
    dm_sluice,
    dm_shifting,
    active_daymen,
    day_idx,
    times,
    constraints,
    prev_day_work,
    min_longest_rest_hours=6,
    max_diff_hours=1.0,
):
    """
    Tasapainottaa daymanien tuntieroja siirtämällä 30 min slotteja.
    Säilyttää op-kattavuuden sekä STCW-vaatimukset.
    """
    if len(active_daymen) < 2:
        return

    max_iterations = 200

    for _ in range(max_iterations):
        hours = {dm: sum(dm_work[dm]) / 2 for dm in active_daymen}
        donor = max(active_daymen, key=lambda dm: hours[dm])
        receiver = min(active_daymen, key=lambda dm: hours[dm])
        diff = hours[donor] - hours[receiver]

        if diff <= max_diff_hours:
            break

        moved = False

        for slot in range(NORMAL_START, NORMAL_END):
            if LUNCH_START <= slot < LUNCH_END:
                continue
            if not dm_work[donor][slot] or dm_work[receiver][slot]:
                continue

            # Älä siirrä pakollisia blokkeja donorilta
            if (
                dm_arr[donor][slot]
                or dm_dep[donor][slot]
                or dm_sluice[donor][slot]
                or dm_shifting[donor][slot]
                or must_work_slot(donor, slot, day_idx, constraints)
            ):
                continue

            donor_hours_after = hours[donor] - 0.5
            if donor_hours_after < get_min_hours(donor, constraints):
                continue

            receiver_hours = hours[receiver]
            receiver_max = get_max_hours(receiver, constraints)
            if receiver_hours >= receiver_max:
                continue
            if not can_work_slot(receiver, slot, day_idx, constraints, receiver_hours):
                continue

            # Op-kattavuus: jos donor on ainoa, receiverin pitää ottaa slotti.
            is_op = is_op_slot(times, slot)
            op_workers = [
                dm for dm in active_daymen
                if dm_work[dm][slot] and dm_ops[dm][slot]
            ]
            donor_is_only_op = is_op and op_workers == [donor]

            # Testaa STCW donorille ja receiverille
            donor_test = dm_work[donor][:]
            donor_test[slot] = False
            receiver_test = dm_work[receiver][:]
            receiver_test[slot] = True

            if not check_stcw_ok(
                donor_test,
                prev_day_work[donor],
                min_longest_rest_hours=min_longest_rest_hours
            ):
                continue
            if not check_stcw_ok(
                receiver_test,
                prev_day_work[receiver],
                min_longest_rest_hours=min_longest_rest_hours
            ):
                continue

            # Tee siirto
            dm_work[donor][slot] = False
            dm_work[receiver][slot] = True

            if is_op:
                dm_ops[receiver][slot] = True
                # donorin op-merkintä pois, jos donor ei enää työskentele slotissa
                dm_ops[donor][slot] = False
            elif dm_ops[donor][slot] and not dm_work[donor][slot]:
                dm_ops[donor][slot] = False

            # Varmista, että op-slotilla on edelleen vähintään yksi tekijä
            if donor_is_only_op and not any(
                dm_work[dm][slot] and dm_ops[dm][slot] for dm in active_daymen
            ):
                # rollback
                dm_work[donor][slot] = True
                dm_work[receiver][slot] = False
                dm_ops[donor][slot] = True
                if not dm_work[receiver][slot]:
                    dm_ops[receiver][slot] = False
                continue

            moved = True
            break

        if not moved:
            break


def fix_stcw_violations(
    dm_work,
    dm_ops,
    dm_arr,
    dm_dep,
    dm_sluice,
    dm_shifting,
    active_daymen,
    day_idx,
    times,
    constraints,
    prev_day_work,
    min_longest_rest_hours=6,
    wm_work=None,
    wm_sluice=None,
    watchman_states=None,
    pending_next_day=None,
):
    """
    VAIHE 6: Korjaa STCW-rikkeet jälkikäteen.
    
    Jos työntekijällä on STCW-rike (alle 10h lepoa tai alle 6h pisin lepo
    jossain 24h ikkunassa), yritetään siirtää slotteja toiselle työntekijälle.
    
    Tarkistaa sekä:
    - prev_day -> current_day (edellinen päivä vaikuttaa nykyiseen)
    - current_day -> next_day (nykyinen päivä vaikuttaa seuraavaan, simuloitu)
    
    Priorisoi:
    1. Slussi-slottien siirto watchmanille (jos mahdollista)
    2. Aamuslottien siirto toiselle daymanille
    """
    if wm_work is None:
        wm_work = {wm: [False] * 48 for wm in WATCHMEN}
    if wm_sluice is None:
        wm_sluice = {wm: [False] * 48 for wm in WATCHMEN}
    if watchman_states is None:
        watchman_states = {wm: {'extended_start': False, 'extended_end': False} for wm in WATCHMEN}
    if pending_next_day is None:
        pending_next_day = {dm: {'work': [], 'sluice': []} for dm in DAYMEN}
    
    def check_stcw_both_directions(dm, test_work):
        """
        Tarkistaa STCW molempiin suuntiin:
        1. prev_day -> test_work (nykyinen)
        2. test_work -> next_day (simuloitu)
        """
        # Tarkista edellinen -> nykyinen
        ok1, worst1, analysis1 = check_stcw_sliding(
            prev_day_work.get(dm, [False] * 48),
            test_work,
            min_longest_rest_hours
        )
        
        if not ok1:
            return False, worst1, analysis1
        
        # Simuloi seuraava päivä: carry-over + normaali 08-16 työpäivä
        next_day_test = [False] * 48
        
        # Lisää carry-over slotit
        for s in pending_next_day.get(dm, {}).get('work', []):
            if 0 <= s < 48:
                next_day_test[s] = True
        
        # Lisää normaali työpäivä (08-16, lounas pois)
        for s in range(NORMAL_START, NORMAL_END):
            if LUNCH_START <= s < LUNCH_END:
                continue
            next_day_test[s] = True
        
        # Tarkista nykyinen -> seuraava
        ok2, worst2, analysis2 = check_stcw_sliding(
            test_work,
            next_day_test,
            min_longest_rest_hours
        )
        
        if not ok2:
            return False, worst2, analysis2
        
        return True, None, analysis1
    
    max_iterations = 100
    
    for iteration in range(max_iterations):
        violation_found = False
        
        for dm in active_daymen:
            ok, worst_slot, analysis = check_stcw_both_directions(dm, dm_work[dm])
            
            if ok:
                continue
            
            violation_found = True
            moved = False
            
            # VAIHE 1: Yritä siirtää slussi-slotteja watchmanille
            sluice_slots = [s for s in range(48) if dm_sluice[dm][s]]
            
            for slot in sluice_slots:
                # Etsi watchman joka voi ottaa slotin
                wm = find_available_watchman(slot, watchman_states, wm_work)
                
                if wm is None:
                    continue
                
                # Testaa siirto
                dm_test = dm_work[dm][:]
                dm_test[slot] = False
                
                dm_ok_after, _, new_analysis = check_stcw_both_directions(dm, dm_test)
                
                # Jos tilanne ei parane, ohita
                if not dm_ok_after:
                    if analysis and new_analysis:
                        old_rest = analysis['longest_rest']
                        new_rest = new_analysis['longest_rest']
                        if new_rest <= old_rest:
                            continue
                    elif not new_analysis:
                        continue
                
                # Tee siirto
                dm_work[dm][slot] = False
                dm_sluice[dm][slot] = False
                wm_work[wm][slot] = True
                wm_sluice[wm][slot] = True
                update_watchman_state(wm, slot, watchman_states)
                
                # Päivitä myös pending_next_day jos slotti oli carry-over
                if slot in pending_next_day.get(dm, {}).get('work', []):
                    pending_next_day[dm]['work'].remove(slot)
                if slot in pending_next_day.get(dm, {}).get('sluice', []):
                    pending_next_day[dm]['sluice'].remove(slot)
                
                moved = True
                break
            
            if moved:
                break  # Aloita uusi iteraatio alusta
            
            # VAIHE 2: Yritä siirtää normaaleja slotteja toiselle daymanille
            morning_slots = list(range(NORMAL_START, LUNCH_START))  # 08:00-11:30
            other_slots = list(range(LUNCH_END, NORMAL_END))  # 12:00-17:00
            
            for slot in morning_slots + other_slots:
                if not dm_work[dm][slot]:
                    continue
                
                # Älä siirrä pakollisia
                if (dm_arr[dm][slot] or dm_dep[dm][slot] or 
                    dm_sluice[dm][slot] or dm_shifting[dm][slot] or
                    must_work_slot(dm, slot, day_idx, constraints)):
                    continue
                
                # Etsi toinen työntekijä joka voi ottaa slotin
                for receiver in active_daymen:
                    if receiver == dm:
                        continue
                    
                    if dm_work[receiver][slot]:
                        continue  # Receiver jo töissä tässä slotissa
                    
                    receiver_hours = sum(dm_work[receiver]) / 2
                    receiver_max = get_max_hours(receiver, constraints)
                    
                    if receiver_hours >= receiver_max:
                        continue
                    
                    if not can_work_slot(receiver, slot, day_idx, constraints, receiver_hours):
                        continue
                    
                    # Testaa siirto
                    dm_test = dm_work[dm][:]
                    dm_test[slot] = False
                    
                    receiver_test = dm_work[receiver][:]
                    receiver_test[slot] = True
                    
                    # Tarkista että siirto parantaa tilannetta
                    dm_ok_after, _, new_analysis = check_stcw_both_directions(dm, dm_test)
                    
                    # Jos tilanne ei parane, ohita
                    if not dm_ok_after:
                        if analysis and new_analysis:
                            old_rest = analysis['total_rest']
                            new_rest = new_analysis['total_rest']
                            if new_rest <= old_rest:
                                continue
                        elif not new_analysis:
                            continue
                    
                    # Tarkista ettei siirto aiheuta rikettä receiverille
                    receiver_ok, _, _ = check_stcw_both_directions(receiver, receiver_test)
                    
                    if not receiver_ok:
                        continue  # Siirto aiheuttaisi rikkeen receiverille
                    
                    # Tarkista donor min tunnit
                    dm_hours_after = sum(dm_test) / 2
                    if dm_hours_after < get_min_hours(dm, constraints):
                        continue
                    
                    # Op-kattavuus tarkistus
                    is_op = is_op_slot(times, slot)
                    
                    # Tee siirto
                    dm_work[dm][slot] = False
                    dm_work[receiver][slot] = True
                    
                    if is_op:
                        dm_ops[dm][slot] = False
                        dm_ops[receiver][slot] = True
                    
                    moved = True
                    break
                
                if moved:
                    break
            
            if moved:
                break  # Aloita uusi iteraatio alusta
        
        if not violation_found:
            break  # Kaikki OK


# ============================================================================
# BOSUN JA WATCHMANIT
# ============================================================================

def generate_bosun_schedule(times):
    """
    Generoi bosunin työvuorot (08-17 + tulo/lähtö + slussi + shiftaus).
    """
    bosun_work = [False] * 48
    bosun_arr = [False] * 48
    bosun_dep = [False] * 48
    bosun_sluice = [False] * 48
    bosun_shifting = [False] * 48
    
    for arrival_start in times.get('arrival_starts', []):
        add_block(bosun_work, arrival_start, arrival_start + 2, bosun_arr)

    for departure_start in times.get('departure_starts', []):
        add_block(bosun_work, departure_start, departure_start + 2, bosun_dep)

    for sluice_arr_start in times.get('sluice_arr_starts', []):
        add_block(bosun_work, sluice_arr_start, sluice_arr_start + 5, bosun_sluice)

    for sluice_dep_start in times.get('sluice_dep_starts', []):
        add_block(bosun_work, sluice_dep_start, sluice_dep_start + 5, bosun_sluice)

    for shifting_start in times.get('shifting_starts', []):
        add_block(bosun_work, shifting_start, shifting_start + 2, bosun_shifting)
    
    for slot in range(NORMAL_START, NORMAL_END):
        if LUNCH_START <= slot < LUNCH_END:
            continue
        bosun_work[slot] = True
    
    return {
        'work_slots': bosun_work,
        'arrival_slots': bosun_arr,
        'departure_slots': bosun_dep,
        'port_op_slots': [False] * 48,
        'sluice_slots': bosun_sluice,
        'shifting_slots': bosun_shifting
    }


def generate_watchman_schedule(worker, wm_work=None, wm_sluice=None):
    """
    Generoi watchmanin 4-on / 8-off -vuoron (8h/vrk).
    Yhdistää mahdolliset slussi-vuorot vakiovuoroihin.

    Watchman 1: 00-04, 12-16
    Watchman 2: 04-08, 16-20
    Watchman 3: 08-12, 20-24
    """
    # Jos slussi-vuoroja on annettu, käytä niitä pohjana
    if wm_work is not None and worker in wm_work:
        work_slots = wm_work[worker][:]
    else:
        work_slots = [False] * 48
    
    if wm_sluice is not None and worker in wm_sluice:
        sluice_slots = wm_sluice[worker][:]
    else:
        sluice_slots = [False] * 48

    # Lisää vakiovuorot
    watch_blocks = {
        'Watchman 1': [(0, 8), (24, 32)],
        'Watchman 2': [(8, 16), (32, 40)],
        'Watchman 3': [(16, 24), (40, 48)],
    }

    for start, end in watch_blocks.get(worker, []):
        add_block(work_slots, start, end)

    return {
        'work_slots': work_slots,
        'arrival_slots': [False] * 48,
        'departure_slots': [False] * 48,
        'port_op_slots': [False] * 48,
        'sluice_slots': sluice_slots,
        'shifting_slots': [False] * 48
    }


def choose_continuous_night_workers(prev_day_daymen_work):
    """
    Valitse jatkuvan yön tekijät:
    - early_worker: se dayman, joka teki edellisen päivän viimeisen iltaslotin
    - late_worker: ensisijaisesti Dayman PH2
    """
    daymen = ['Dayman EU', 'Dayman PH1', 'Dayman PH2']

    latest_worker = None
    latest_slot = -1
    for dm in daymen:
        work = prev_day_daymen_work.get(dm, [False] * 48)
        for slot in range(47, -1, -1):
            if work[slot]:
                if slot > latest_slot:
                    latest_slot = slot
                    latest_worker = dm
                break

    early_worker = latest_worker or 'Dayman EU'
    late_worker = 'Dayman PH2'
    if late_worker == early_worker:
        late_worker = 'Dayman PH1'

    return early_worker, late_worker


# ============================================================================
# PÄÄFUNKTIO
# ============================================================================

def generate_schedule(days_data, constraints=None, min_longest_rest_hours=6):
    """
    Generoi työvuorot blokkipohjaisella lähestymistavalla.
    
    Args:
        days_data: Lista päivien tiedoista
        constraints: Lista rajoitteista (valinnainen)
    
    Returns:
        (workbook, all_days, report)
    """
    if constraints is None:
        constraints = []
    
    all_days = {w: [] for w in WORKERS}
    num_days = len(days_data)
    
    # VAIHE 0: Analysoi jatkuvat yövuorot
    continuous_nights = analyze_continuous_nights(days_data)
    
    # Pending-slotit edelliseltä päivältä (carry-over keskiyön yli)
    pending_next_day = {dm: {'work': [], 'sluice': []} for dm in DAYMEN}
    
    # Generoi päivä kerrallaan
    for day_idx, info in enumerate(days_data):
        # Parsitaan päivän ajat
        times = parse_day_times(info)
        
        # Edellisen päivän työvuorot
        prev_day_work = {}
        for dm in DAYMEN:
            if day_idx > 0:
                prev_day_work[dm] = all_days[dm][day_idx - 1]['work_slots']
            else:
                prev_day_work[dm] = [False] * 48
        
        # Tarkista jatkuva yö
        continuous_night_info = None
        for night_info in continuous_nights:
            if night_info['day_index'] == day_idx - 1:
                prev_day_daymen_work = {
                    dm: all_days[dm][day_idx - 1]['work_slots']
                    for dm in DAYMEN
                }
                early_worker, late_worker = choose_continuous_night_workers(prev_day_daymen_work)

                split_slot = time_to_slot(1, 0)

                continuous_night_info = {
                    'early_worker': early_worker,
                    'late_worker': late_worker,
                    'split_slot': split_slot
                }
                break
        
        # Alusta työvuorolistat
        dm_work = {dm: [False] * 48 for dm in DAYMEN}
        dm_arr = {dm: [False] * 48 for dm in DAYMEN}
        dm_dep = {dm: [False] * 48 for dm in DAYMEN}
        dm_ops = {dm: [False] * 48 for dm in DAYMEN}
        dm_sluice = {dm: [False] * 48 for dm in DAYMEN}
        dm_shifting = {dm: [False] * 48 for dm in DAYMEN}
        
        # Alusta watchman-tietorakenteet
        wm_work = {wm: [False] * 48 for wm in WATCHMEN}
        wm_sluice = {wm: [False] * 48 for wm in WATCHMEN}
        watchman_states = {wm: {'extended_start': False, 'extended_end': False} for wm in WATCHMEN}
        
        # LISÄÄ PENDING-SLOTIT EDELLISELTÄ PÄIVÄLTÄ (carry-over)
        for dm in DAYMEN:
            for slot in pending_next_day[dm]['work']:
                if 0 <= slot < 48:
                    dm_work[dm][slot] = True
            for slot in pending_next_day[dm]['sluice']:
                if 0 <= slot < 48:
                    dm_sluice[dm][slot] = True
        
        # Nollaa pending seuraavaa päivää varten
        pending_next_day = {dm: {'work': [], 'sluice': []} for dm in DAYMEN}
        
        # Aktiiviset daymanit
        active_daymen = [dm for dm in DAYMEN if not is_day_off(dm, day_idx, constraints)]
        
        # VAIHE 0.5: Pakolliset slotit rajoitteista
        apply_constraint_slots(dm_work, dm_ops, DAYMEN, day_idx, times, constraints)
        
        # VAIHE 1: Pakolliset
        apply_arrival_slots(dm_work, dm_arr, active_daymen, day_idx, times, constraints)
        apply_departure_slots(dm_work, dm_dep, active_daymen, day_idx, times, constraints)
        pending_next_day = apply_sluice_arrival_slots(
            dm_work, dm_sluice, DAYMEN, times, pending_next_day,
            prev_day_work=prev_day_work, wm_work=wm_work, wm_sluice=wm_sluice,
            watchman_states=watchman_states, min_longest_rest_hours=min_longest_rest_hours
        )
        pending_next_day = apply_sluice_departure_slots(
            dm_work, dm_sluice, DAYMEN, times, pending_next_day,
            prev_day_work=prev_day_work, wm_work=wm_work, wm_sluice=wm_sluice,
            watchman_states=watchman_states, min_longest_rest_hours=min_longest_rest_hours
        )
        apply_shifting_slots(dm_work, dm_shifting, DAYMEN, times)
        apply_op_outside_normal_hours(
            dm_work, dm_ops, active_daymen, day_idx, times,
            constraints, prev_day_work, continuous_night_info,
            min_longest_rest_hours=min_longest_rest_hours
        )
        
        # VAIHE 3: Jaa työblokit
        op_inside_slots = fill_op_inside_normal_hours(
            dm_work, dm_ops, active_daymen, day_idx, times, constraints, prev_day_work,
            min_longest_rest_hours=min_longest_rest_hours
        )
        fill_remaining_hours(dm_work, dm_ops, active_daymen, day_idx, times, constraints,
                            prev_day_work=prev_day_work, min_longest_rest_hours=min_longest_rest_hours)
        ensure_op_coverage(dm_work, dm_ops, op_inside_slots, active_daymen, day_idx, constraints)
        fill_gaps_between_blocks(dm_work, dm_ops, active_daymen, day_idx, times, constraints)

        # VAIHE 5: Tasapainota daymanien tunnit (ero max 1h)
        rebalance_dayman_hours(
            dm_work, dm_ops, dm_arr, dm_dep, dm_sluice, dm_shifting,
            active_daymen, day_idx, times, constraints, prev_day_work,
            min_longest_rest_hours=min_longest_rest_hours,
            max_diff_hours=1.0
        )
        
        # VAIHE 6: Korjaa STCW-rikkeet
        fix_stcw_violations(
            dm_work, dm_ops, dm_arr, dm_dep, dm_sluice, dm_shifting,
            active_daymen, day_idx, times, constraints, prev_day_work,
            min_longest_rest_hours=min_longest_rest_hours,
            wm_work=wm_work, wm_sluice=wm_sluice, watchman_states=watchman_states,
            pending_next_day=pending_next_day
        )
        
        # VAIHE 7: Täytä pienet aukot (max 2h) - kutsutaan lopuksi kun kaikki muu on valmis
        fill_small_gaps(dm_work, dm_ops, active_daymen, times)
        
        # Tallenna daymanien tulokset
        for dm in DAYMEN:
            all_days[dm].append({
                'work_slots': dm_work[dm],
                'arrival_slots': dm_arr[dm],
                'departure_slots': dm_dep[dm],
                'port_op_slots': dm_ops[dm],
                'sluice_slots': dm_sluice[dm],
                'shifting_slots': dm_shifting[dm]
            })
        
        # Bosun
        all_days['Bosun'].append(generate_bosun_schedule(times))
        
        # Watchmanit (yhdistää slussi-vuorot vakiovuoroihin)
        for wm in WATCHMEN:
            all_days[wm].append(generate_watchman_schedule(wm, wm_work, wm_sluice))
    
    # Rakenna Excel
    wb, report = build_workbook_and_report(all_days, num_days, WORKERS)
    
    return wb, all_days, report


# ============================================================================
# EXCEL-GENEROINTI
# ============================================================================

def build_workbook_and_report(all_days, num_days, workers):
    """
    Rakentaa Excel-työkirjan ja raportin.
    """
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True
    ws = wb.active
    ws.title = "Työvuorot"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    report_lines = []
    current_row = 1
    
    for d in range(num_days):
        ws.cell(row=current_row, column=1, value=f"Päivä {d+1}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Työntekijä")
        for slot in range(48):
            col = slot + 2
            # Näytä vain tasatunnit (slotit 0, 2, 4, 6, ... = 00:00, 01:00, 02:00, ...)
            if slot % 2 == 0:
                time_str = f"{slot // 2:02d}:00"
                ws.cell(row=current_row, column=col, value=time_str)
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=col).font = Font(size=8)
        ws.cell(row=current_row, column=50, value="Tunnit")
        current_row += 1

        worker_start_row = current_row
        for worker in workers:
            ws.cell(row=current_row, column=1, value=worker)
            day_data = all_days[worker][d]
            work = day_data['work_slots']
            arr = day_data['arrival_slots']
            dep = day_data['departure_slots']
            ops = day_data.get('port_op_slots', [False] * 48)
            sluice = day_data.get('sluice_slots', [False] * 48)
            shifting = day_data.get('shifting_slots', [False] * 48)
            
            hours = sum(work) / 2
            
            for slot in range(48):
                col = slot + 2
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                
                if sluice[slot]:
                    cell.value = "SL"
                elif shifting[slot]:
                    cell.value = "SH"
                elif arr[slot]:
                    cell.value = "B"
                elif dep[slot]:
                    cell.value = "C"
                elif ops[slot]:
                    cell.value = "OP"
                elif work[slot]:
                    cell.value = "X"
                else:
                    cell.value = None
                
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=8)
            
            hours_cell = ws.cell(row=current_row, column=50)
            hours_cell.value = f'=COUNTA(B{current_row}:AW{current_row})/2'
            hours_cell.number_format = '0.0'
            current_row += 1
            
            ranges = get_work_ranges(work)
            report_lines.append(f"Päivä {d+1} - {worker}: {hours}h | {' + '.join(ranges)}")

        add_time_cell_conditional_formatting(ws, worker_start_row, current_row - 1)
        current_row += 1
    
    ws.column_dimensions['A'].width = 12
    for col in range(2, 50):
        ws.column_dimensions[get_column_letter(col)].width = 4
    ws.column_dimensions[get_column_letter(50)].width = 6
    
    report = "\n".join(report_lines)
    return wb, report


# ============================================================================
# MANUAALINEN PÄIVÄ 1
# ============================================================================

def generate_schedule_with_manual_day1(days_data, manual_day1_slots, min_longest_rest_hours=6):
    """
    Generoi työvuorot manuaalisella päivällä 1.
    """
    if not days_data:
        return generate_schedule(days_data)
    
    all_days = {w: [] for w in WORKERS}
    num_days = len(days_data)
    
    for worker in WORKERS:
        if worker in manual_day1_slots:
            work = manual_day1_slots[worker]
        else:
            work = [False] * 48
        
        all_days[worker].append({
            'work_slots': work,
            'arrival_slots': [False] * 48,
            'departure_slots': [False] * 48,
            'port_op_slots': [False] * 48,
            'sluice_slots': [False] * 48,
            'shifting_slots': [False] * 48
        })
    
    if num_days > 1:
        _, rest_days, _ = generate_schedule(
            days_data[1:],
            min_longest_rest_hours=min_longest_rest_hours
        )
        for worker in WORKERS:
            all_days[worker].extend(rest_days[worker])
    
    wb, report = build_workbook_and_report(all_days, num_days, WORKERS)
    return wb, all_days, report
